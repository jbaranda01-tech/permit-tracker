import os
import re
import uuid
import unicodedata
import secrets
import click
import logging
from datetime import datetime, date, timedelta
from decimal import Decimal, InvalidOperation
from functools import wraps

from io import BytesIO
from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, send_file, jsonify, abort, session
)
from flask_login import (
    LoginManager, login_user, logout_user,
    login_required, current_user
)
from flask_migrate import Migrate
from werkzeug.utils import secure_filename

import hashlib
from config import Config
from flask_compress import Compress
from models import (
    db, User, Employee, EmployeePermit, Equipment, EquipmentPermit,
    CompanyPermit, EMPLOYEE_PERMIT_TYPES, EQUIPMENT_PERMIT_TYPES,
    COMPANY_PERMIT_TYPES, FileStorage, NotificationLog,
    protect_uploaded_files, ProtectedFileError,
    Issue, IssueStatusHistory, IssuePhoto, UserIssueRole,
    ISSUE_CATEGORIES, ISSUE_SEVERITIES, ISSUE_STATUSES,
    RESOLVED_STATUSES, LEGACY_ISSUE_STATUSES,
    EMPLOYEE_ARCHIVE_REASONS, EQUIPMENT_ARCHIVE_REASONS,
    EQUIPMENT_CLASSES, classify_equipment,
    INSURANCE_TYPE_BY_CLASS,
    INSURANCE_OWN, INSURANCE_POLICY_CHOICES, sync_vehicle_insurance_permit,
)
from list_prefs import (
    DASHBOARD_PREF_PARAMS, restore_list_args, remember_list_args,
    forget_list_args,
)

# ── APP INIT ───────────────────────────────────────────────────────────

app = Flask(__name__)
app.config.from_object(Config)
Compress(app)

# Ensure logs go to stdout for Railway
if not app.debug:
    gunicorn_logger = logging.getLogger('gunicorn.error')
    app.logger.handlers = gunicorn_logger.handlers
    app.logger.setLevel(gunicorn_logger.level or logging.INFO)

db.init_app(app)
migrate = Migrate(app, db)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Por favor inicie sesión para acceder.'
login_manager.login_message_category = 'warning'

from issues import issues_bp
app.register_blueprint(issues_bp)


ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'doc', 'docx'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def save_uploaded_file(file):
    """Save an uploaded file to the database and return the storage filename."""
    filename = f"{uuid.uuid4().hex}_{secure_filename(file.filename)}"
    file_data = file.read()
    stored = FileStorage(
        filename=filename,
        original_filename=file.filename,
        mime_type=file.content_type or 'application/octet-stream',
        data=file_data,
        size=len(file_data),
    )
    db.session.add(stored)
    return filename


def delete_stored_file(filename):
    """Delete a file from the database by filename."""
    stored = FileStorage.query.filter_by(filename=filename).first()
    if stored:
        db.session.delete(stored)


# ── NOTIFICATION HELPERS ──────────────────────────────────────────────

def _gather_expiring_items(employee_id=None):
    today = date.today()
    alert_date = today + timedelta(days=app.config['ALERT_DAYS_BEFORE'])

    query = Employee.query.filter(
        Employee.status == 'activo',
        Employee.archived_at.is_(None),
        Employee.email.isnot(None),
        Employee.email != '',
    )
    if employee_id:
        query = query.filter(Employee.id == employee_id)

    results = {}
    for emp in query.all():
        items = []
        if emp.license_expiration:
            if emp.license_expiration < today:
                items.append({'key': f'license:{emp.id}', 'name': 'Licencia de Conducir',
                              'date': emp.license_expiration, 'status': 'expired'})
            elif emp.license_expiration <= alert_date:
                items.append({'key': f'license:{emp.id}', 'name': 'Licencia de Conducir',
                              'date': emp.license_expiration, 'status': 'expiring_soon'})

        for permit in emp.permits:
            if permit.applicability == 'N/A' or permit.expiration_date is None:
                continue
            if permit.expiration_date < today:
                items.append({'key': f'employee_permit:{permit.id}', 'name': permit.display_name,
                              'date': permit.expiration_date, 'status': 'expired'})
            elif permit.expiration_date <= alert_date:
                items.append({'key': f'employee_permit:{permit.id}', 'name': permit.display_name,
                              'date': permit.expiration_date, 'status': 'expiring_soon'})

        if items:
            results[emp.id] = {'employee': emp, 'items': items}
    return results


def _filter_already_notified(employee_id, items):
    cutoff = datetime.utcnow() - timedelta(days=7)
    recent = NotificationLog.query.filter(
        NotificationLog.employee_id == employee_id,
        NotificationLog.sent_at > cutoff,
        NotificationLog.status == 'sent',
    ).all()
    notified_keys = {log.permit_key for log in recent}
    return [item for item in items if item['key'] not in notified_keys]


def send_notification_email(employee, items, dry_run=False):
    if dry_run:
        return (True, None)
    try:
        from sendgrid import SendGridAPIClient
        from sendgrid.helpers.mail import Mail, HtmlContent

        html = render_template('email/permit_expiry_notice.html',
                               employee=employee, items=items)
        message = Mail(
            from_email=app.config['SENDGRID_FROM_EMAIL'],
            to_emails=employee.email,
            subject='Aviso: Documentos por vencer o vencidos',
            html_content=HtmlContent(html),
        )
        sg = SendGridAPIClient(app.config['SENDGRID_API_KEY'])
        response = sg.send(message)
        if response.status_code in (200, 201, 202):
            return (True, None)
        return (False, f'SendGrid status {response.status_code}')
    except Exception as e:
        return (False, str(e))


def run_notification_cycle(dry_run=False, employee_id=None):
    data = _gather_expiring_items(employee_id)
    sent = 0
    skipped = 0
    failed = 0
    details = []

    for emp_id, entry in data.items():
        emp = entry['employee']
        remaining = _filter_already_notified(emp_id, entry['items'])
        if not remaining:
            skipped += 1
            details.append(f'  SKIP {emp.name} — already notified')
            continue

        success, error = send_notification_email(emp, remaining, dry_run=dry_run)

        for item in remaining:
            log = NotificationLog(
                employee_id=emp_id,
                permit_key=item['key'],
                email_to=emp.email,
                status='sent' if success else 'failed',
                error_message=error,
            )
            if not dry_run:
                db.session.add(log)

        if success:
            sent += 1
            details.append(f'  {"[DRY RUN] " if dry_run else ""}SENT {emp.name} ({emp.email}) — {len(remaining)} items')
        else:
            failed += 1
            details.append(f'  FAIL {emp.name} ({emp.email}) — {error}')

    if not dry_run:
        db.session.commit()

    return {'sent': sent, 'skipped': skipped, 'failed': failed, 'details': details}


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


# ── ROLE DECORATORS ────────────────────────────────────────────────────

def manager_required(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not current_user.is_manager:
            flash('No tiene permisos para esta acción.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not current_user.is_admin:
            flash('Solo administradores pueden acceder.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated


# ── TEMPLATE HELPERS ───────────────────────────────────────────────────

@app.template_filter('format_phone')
def format_phone(value):
    if not value:
        return value
    digits = ''.join(c for c in str(value) if c.isdigit())
    if len(digits) == 10:
        return f'({digits[:3]}) {digits[3:6]}-{digits[6:]}'
    if len(digits) == 7:
        return f'{digits[:3]}-{digits[3:]}'
    return value

@app.template_filter('time_ago')
def time_ago(value):
    """Spanish relative time for issue timestamps (naive UTC). Falls back to
    an absolute date beyond ~7 days."""
    if not value:
        return ''
    delta = datetime.utcnow() - value
    seconds = delta.total_seconds()
    if seconds < 0:
        return value.strftime('%m/%d/%Y')
    minutes = int(seconds // 60)
    hours = int(seconds // 3600)
    days = delta.days
    if seconds < 60:
        return 'hace un momento'
    if minutes < 60:
        return f'hace {minutes}m'
    if hours < 24:
        return f'hace {hours}h'
    if days == 1:
        return 'ayer'
    if days < 7:
        return f'hace {days}d'
    return value.strftime('%m/%d/%Y')

@app.context_processor
def inject_globals():
    today = date.today()
    alert_date = today + timedelta(days=30)
    return {
        'today': today,
        'alert_date': alert_date,
    }


# ── ASSET VERSIONING ──────────────────────────────────────────────────

@app.context_processor
def asset_hash():
    def versioned_static(filename):
        filepath = os.path.join(app.static_folder, filename)
        try:
            mtime = str(os.path.getmtime(filepath))
            h = hashlib.md5(mtime.encode()).hexdigest()[:8]
        except OSError:
            h = ''
        return url_for('static', filename=filename) + '?v=' + h
    return {'versioned_static': versioned_static}


# ── REQUEST LOGGING ────────────────────────────────────────────────────

@app.after_request
def log_request(response):
    app.logger.info(f"{request.method} {request.path} → {response.status_code}")
    if request.path.startswith('/static/'):
        response.headers['Cache-Control'] = 'public, max-age=604800'
    return response


# ── MODULE ACCESS ENFORCEMENT ─────────────────────────────────────────

SHOP_EXEMPT_PREFIXES = ('issues.', 'static')
SHOP_EXEMPT_ENDPOINTS = {'login', 'logout', 'health'}

@app.before_request
def enforce_module_access():
    if not current_user.is_authenticated:
        return

    endpoint = request.endpoint
    if endpoint is None:
        return

    if endpoint in SHOP_EXEMPT_ENDPOINTS:
        return

    if any(endpoint.startswith(p) for p in SHOP_EXEMPT_PREFIXES):
        return

    if current_user.is_shop_only:
        flash('Su cuenta solo tiene acceso al módulo de taller.', 'warning')
        return redirect(url_for('issues.queue'))


# ── PWA ────────────────────────────────────────────────────────────────

@app.route('/manifest.json')
def manifest():
    return app.send_static_file('manifest.json')


# ── AUTH ROUTES ────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        if current_user.is_shop_only:
            return redirect(url_for('issues.queue'))
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            login_user(user, remember=True)
            next_page = request.args.get('next')
            flash(f'Bienvenido, {user.username}!', 'success')
            if next_page:
                return redirect(next_page)
            if user.is_shop_only:
                return redirect(url_for('issues.queue'))
            return redirect(url_for('dashboard'))
        flash('Usuario o contraseña incorrectos.', 'danger')
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Sesión cerrada.', 'info')
    return redirect(url_for('login'))


# ── DASHBOARD ──────────────────────────────────────────────────────────

DASHBOARD_SORTS = ('name', 'category', 'urgency', 'expiration')
ATTENTION_STATUSES = ('expired', 'expiring_soon', 'missing')
PERMIT_STATUS_MAP = {'expired': 'expired', 'expiring': 'expiring_soon'}  # url value -> summary key
SPANISH_MONTHS = ('Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                  'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre')


def _exp_month_options(view):
    """('YYYY-MM', 'Agosto 2026') pairs for every month with a counted expiration.

    Sources mirror _expires_in_month: per-entity permits (non-archived owner,
    non-N/A) plus the driver's license on the employees view.
    """
    if view == 'equipment':
        queries = [db.session.query(EquipmentPermit.expiration_date)
                   .join(Equipment)
                   .filter(Equipment.archived_at.is_(None),
                           EquipmentPermit.applicability != 'N/A',
                           EquipmentPermit.expiration_date.isnot(None))]
    else:
        queries = [db.session.query(EmployeePermit.expiration_date)
                   .join(Employee)
                   .filter(Employee.archived_at.is_(None),
                           EmployeePermit.applicability != 'N/A',
                           EmployeePermit.expiration_date.isnot(None)),
                   db.session.query(Employee.license_expiration)
                   .filter(Employee.archived_at.is_(None),
                           Employee.license_expiration.isnot(None))]
    months = {(d.year, d.month) for q in queries for (d,) in q.distinct()}
    return [(f'{y:04d}-{m:02d}', f'{SPANISH_MONTHS[m - 1]} {y}') for y, m in sorted(months)]


def _expires_in_month(item, exp_month, permit_type, view):
    """True if a counted expiration (the given type's, else any) falls in exp_month ('YYYY-MM')."""
    def in_month(d):
        return d is not None and f'{d.year:04d}-{d.month:02d}' == exp_month

    if permit_type:
        if view == 'employees' and permit_type == 'LICENCIA':
            return in_month(item.license_expiration)
        permit = next((p for p in item.permits if p.permit_type == permit_type), None)
        return permit is not None and permit.applicability != 'N/A' and in_month(permit.expiration_date)
    if view == 'employees' and in_month(item.license_expiration):
        return True
    return any(p.applicability != 'N/A' and in_month(p.expiration_date) for p in item.permits)


def _permit_type_status(item, permit_type, view):
    """Status ('expired'|'expiring_soon'|'valid'|'missing'|'na') of one permit type on an entity."""
    if view == 'employees' and permit_type == 'LICENCIA':
        if not item.license_expiration:
            return 'missing'
        today = date.today()
        if item.license_expiration < today:
            return 'expired'
        if item.license_expiration <= today + timedelta(days=Config.ALERT_DAYS_BEFORE):
            return 'expiring_soon'
        return 'valid'
    permit = next((p for p in item.permits if p.permit_type == permit_type), None)
    return permit.status if permit else 'missing'


def _filter_dashboard_items(items, permit_type, permit_status, view):
    """permit_type alone: that type needs attention; with permit_status: that type has exactly
    that status; permit_status alone: any permit has that status."""
    if permit_type:
        if permit_status:
            wanted = PERMIT_STATUS_MAP[permit_status]
            return [i for i in items if _permit_type_status(i, permit_type, view) == wanted]
        return [i for i in items if _permit_type_status(i, permit_type, view) in ATTENTION_STATUSES]
    if permit_status:
        wanted = PERMIT_STATUS_MAP[permit_status]
        return [i for i in items if i.permit_status_summary[wanted] > 0]
    return items


def _sort_dashboard_items(items, sort_by, view, permit_type=''):
    def entity_name(item):
        return (item.name if view == 'employees' else item.display_name).lower()

    if sort_by == 'category':
        def key(item):
            value = ((item.area if view == 'employees' else item.model) or '').strip().lower()
            return (value == '', value, entity_name(item))
        return sorted(items, key=key)
    if sort_by == 'urgency':
        def key(item):
            s = item.permit_status_summary
            if s['expired'] > 0:
                rank = 0
            elif s['expiring_soon'] > 0:
                rank = 1
            elif s['missing'] > 0:
                rank = 2
            else:
                rank = 3
            return (rank, -s['expired'], -s['expiring_soon'], entity_name(item))
        return sorted(items, key=key)
    if sort_by == 'expiration':
        if permit_type:
            # Sort by the selected type's own expiration date: already-expired
            # first (ascending), soonest upcoming next, N/A/missing/absent last.
            def key(item):
                if view == 'employees' and permit_type == 'LICENCIA':
                    d = item.license_expiration
                else:
                    p = next((p for p in item.permits if p.permit_type == permit_type), None)
                    d = p.expiration_date if p and p.applicability != 'N/A' else None
                return (d or date.max, entity_name(item))
            return sorted(items, key=key)
        return sorted(items, key=lambda i: (i.next_expiration or date.max, entity_name(i)))
    return sorted(items, key=entity_name)


# ── DASHBOARD LIST CONTEXT ─────────────────────────────────────────────

def _dashboard_context(args, view=None, with_counts=True):
    """Normalized params + the filtered/sorted LB/PLI/Personal panels for a set
    of dashboard GET args.

    Shared by dashboard() and _sibling_nav() so the sequence the detail-page
    arrows walk can never drift from the list on screen. `view` overrides
    args['view'] — a detail page knows which list it belongs to. The 'company'
    view is handled by the route; this helper only knows employees/equipment.
    `with_counts=False` skips the tile totals, whose per-item
    `permit_status_summary` loads every permit row the nav doesn't need.
    """
    view = view or args.get('view', 'employees')
    if view not in ('employees', 'equipment'):
        view = 'employees'
    # A request with no list state of its own (sidebar click, post-edit
    # redirect) picks up this user's remembered filters/sort for THIS view.
    # Keyed per view: permit_type/equipment_class/exp_month are view-scoped.
    args = restore_list_args(f'dashboard:{view}', args, DASHBOARD_PREF_PARAMS)
    search = args.get('search', '').strip()
    company_filter = args.get('company', '')
    status_filter = args.get('status', '')
    sort_by = args.get('sort', 'name')
    if sort_by not in DASHBOARD_SORTS:
        sort_by = 'name'
    permit_status = args.get('permit_status', '')
    if permit_status not in PERMIT_STATUS_MAP:
        permit_status = ''
    if view == 'equipment':
        permit_types = list(EQUIPMENT_PERMIT_TYPES)
    else:
        permit_types = [('LICENCIA', 'Licencia de Conducir')] + list(EMPLOYEE_PERMIT_TYPES)
    permit_type = args.get('permit_type', '')
    if permit_type not in {code for code, _label in permit_types}:
        permit_type = ''
    equipment_class = args.get('equipment_class', '')
    if view != 'equipment' or equipment_class not in {code for code, _ in EQUIPMENT_CLASSES}:
        equipment_class = ''

    exp_month_options = _exp_month_options(view)
    exp_month = args.get('exp_month', '')
    if exp_month not in {value for value, _label in exp_month_options}:
        exp_month = ''

    if view == 'equipment':
        query = Equipment.query.filter(Equipment.archived_at.is_(None))
        if search:
            query = query.filter(
                db.or_(
                    Equipment.unit_number.ilike(f'%{search}%'),
                    Equipment.plate_number.ilike(f'%{search}%'),
                    Equipment.make.ilike(f'%{search}%'),
                    Equipment.model.ilike(f'%{search}%'),
                    Equipment.vin_serial.ilike(f'%{search}%'),
                )
            )
        if company_filter:
            query = query.filter(Equipment.company == company_filter)
        if status_filter:
            query = query.filter(Equipment.status == status_filter)
        if equipment_class:
            query = query.filter(Equipment.equipment_class == equipment_class)

        # Split by company
        lb_items = query.filter(Equipment.company == 'LB').all()
        pli_items = query.filter(Equipment.company == 'PLI').all()
        personal_items = query.filter(Equipment.company == 'Personal').all()
    else:
        query = Employee.query.filter(Employee.archived_at.is_(None))
        if search:
            query = query.filter(
                db.or_(
                    Employee.name.ilike(f'%{search}%'),
                    Employee.license_number.ilike(f'%{search}%'),
                    Employee.area.ilike(f'%{search}%'),
                )
            )
        if company_filter:
            query = query.filter(Employee.company == company_filter)
        if status_filter:
            query = query.filter(Employee.status == status_filter)

        # Split by company
        lb_items = query.filter(Employee.company == 'LB').all()
        pli_items = query.filter(Employee.company == 'PLI').all()
        personal_items = []

    # Scope filters first (exp_month replaces the type "needs attention" rule when
    # set), then tile counts (stable regardless of the active permit_status tile),
    # then the permit_status filter + sort.
    if exp_month:
        lb_items = [i for i in lb_items if _expires_in_month(i, exp_month, permit_type, view)]
        pli_items = [i for i in pli_items if _expires_in_month(i, exp_month, permit_type, view)]
        personal_items = [i for i in personal_items if _expires_in_month(i, exp_month, permit_type, view)]
    else:
        lb_items = _filter_dashboard_items(lb_items, permit_type, '', view)
        pli_items = _filter_dashboard_items(pli_items, permit_type, '', view)
        personal_items = _filter_dashboard_items(personal_items, permit_type, '', view)

    all_items = lb_items + pli_items + personal_items
    tile_counts = {'expired': 0, 'expiring': 0, 'total': len(all_items)}
    if with_counts:
        for item in all_items:
            summary = item.permit_status_summary
            tile_counts['expired'] += summary['expired']
            tile_counts['expiring'] += summary['expiring_soon']

    if permit_status:
        lb_items = _filter_dashboard_items(lb_items, permit_type, permit_status, view)
        pli_items = _filter_dashboard_items(pli_items, permit_type, permit_status, view)
        personal_items = _filter_dashboard_items(personal_items, permit_type, permit_status, view)

    lb_items = _sort_dashboard_items(lb_items, sort_by, view, permit_type)
    pli_items = _sort_dashboard_items(pli_items, sort_by, view, permit_type)
    personal_items = _sort_dashboard_items(personal_items, sort_by, view, permit_type)

    # Drives the "Limpiar filtros" link only. Unlike the queue's flag of the
    # same name, sort counts: sort is sticky now, so it must be resettable.
    filters_active = bool(search or company_filter or status_filter
                          or permit_type or permit_status or exp_month
                          or equipment_class or sort_by != 'name')

    # The params that identify "which list am I browsing": they ride from the
    # dashboard into every entity link so a detail page can rebuild this exact
    # sequence for its prev/next arrows. Only non-default values ride along, so
    # URLs stay clean (the `... or none` idiom, on the normalized values).
    # A new list param belongs here, or the arrows walk an unfiltered list.
    nav_args = {key: value for key, value in (
        ('view', view if view != 'employees' else ''),
        ('search', search),
        ('company', company_filter),
        ('status', status_filter),
        ('sort', sort_by if sort_by != 'name' else ''),
        ('permit_type', permit_type),
        ('permit_status', permit_status),
        ('exp_month', exp_month),
        ('equipment_class', equipment_class),
    ) if value}

    return {
        'view': view,
        'lb_items': lb_items,
        'pli_items': pli_items,
        'personal_items': personal_items,
        'search': search,
        'company_filter': company_filter,
        'status_filter': status_filter,
        'sort_by': sort_by,
        'permit_type': permit_type,
        'permit_status': permit_status,
        'permit_types': permit_types,
        'exp_month': exp_month,
        'exp_months': exp_month_options,
        'equipment_class': equipment_class,
        'equipment_classes': EQUIPMENT_CLASSES,
        'tile_counts': tile_counts,
        'filters_active': filters_active,
        'nav_args': nav_args,
    }


def _sibling_nav(view, current_id, args):
    """Prev/next links for an entity detail page.

    Walks the same filtered + sorted sequence the dashboard rendered, flattened
    LB → PLI → Personal, so the arrows follow exactly what was on screen.
    Returns None when the record isn't in that list (archived, or filtered out
    by the args it was reached with) — the template then renders no arrows.
    """
    ctx = _dashboard_context(args, view=view, with_counts=False)
    items = ctx['lb_items'] + ctx['pli_items'] + ctx['personal_items']
    ids = [item.id for item in items]
    if current_id not in ids:
        return None
    index = ids.index(current_id)
    endpoint = 'equipment_detail' if view == 'equipment' else 'employee_detail'
    nav_args = ctx['nav_args']

    def entry(item):
        if item is None:
            return None
        label = item.display_name if view == 'equipment' else item.name
        return {'url': url_for(endpoint, id=item.id, **nav_args), 'label': label}

    return {
        'prev': entry(items[index - 1] if index > 0 else None),
        'next': entry(items[index + 1] if index + 1 < len(items) else None),
        'position': index + 1,
        'total': len(items),
        'list_url': url_for('dashboard', **nav_args),
    }


@app.route('/')
@login_required
def dashboard():
    view = request.args.get('view', 'employees')  # employees, equipment or company

    # "Limpiar filtros" — forget this view's remembered choices, then land on a
    # clean URL (which restores nothing, the bucket having just been dropped).
    if request.args.get('reset'):
        if view in ('employees', 'equipment'):
            forget_list_args(f'dashboard:{view}')
        return redirect(url_for('dashboard', view=view if view != 'employees' else None))

    if view == 'company':
        # Backfill missing company permits
        changed = False
        for code, name, companies in COMPANY_PERMIT_TYPES:
            for company in companies:
                existing = CompanyPermit.query.filter_by(company=company, permit_type=code).first()
                if not existing:
                    db.session.add(CompanyPermit(company=company, permit_type=code, applicability='YES'))
                    changed = True
        if changed:
            db.session.commit()

        lb_permits = CompanyPermit.query.filter_by(company='LB').order_by(CompanyPermit.permit_type).all()
        pli_permits = CompanyPermit.query.filter_by(company='PLI').order_by(CompanyPermit.permit_type).all()

        return render_template('dashboard.html',
            view=view,
            lb_permits=lb_permits,
            pli_permits=pli_permits,
        )

    context = _dashboard_context(request.args)
    remember_list_args(f'dashboard:{context["view"]}', context['nav_args'],
                       DASHBOARD_PREF_PARAMS)
    return render_template('dashboard.html', **context)


@app.route('/archive')
@login_required
def archive_view():
    search = request.args.get('search', '').strip()
    emp_q = Employee.query.filter(Employee.archived_at.isnot(None))
    eq_q = Equipment.query.filter(Equipment.archived_at.isnot(None))
    if search:
        emp_q = emp_q.filter(Employee.name.ilike(f'%{search}%'))
        eq_q = eq_q.filter(db.or_(
            Equipment.unit_number.ilike(f'%{search}%'),
            Equipment.plate_number.ilike(f'%{search}%'),
            Equipment.vin_serial.ilike(f'%{search}%'),
            Equipment.make.ilike(f'%{search}%'),
            Equipment.model.ilike(f'%{search}%'),
        ))
    archived_employees = emp_q.order_by(Employee.archived_at.desc()).all()
    archived_equipment = eq_q.order_by(Equipment.archived_at.desc()).all()
    return render_template('archive.html',
                           archived_employees=archived_employees,
                           archived_equipment=archived_equipment,
                           search=search)


# ── EMPLOYEE ROUTES ────────────────────────────────────────────────────

@app.route('/employee/<int:id>')
@login_required
def employee_detail(id):
    emp = Employee.query.get_or_404(id)

    # Backfill missing permit slots for existing employees (skip archived —
    # frozen records shouldn't gain permit rows on a GET)
    if not emp.is_archived:
        existing_types = {}
        for p in emp.permits.all():
            existing_types[p.permit_type] = p
        changed = False
        for code, name in EMPLOYEE_PERMIT_TYPES:
            if code == 'OTHER':
                continue
            if code not in existing_types:
                db.session.add(EmployeePermit(
                    employee_id=emp.id,
                    permit_type=code,
                    applicability='YES' if code == 'PRIMEROS_AUXILIOS' else 'N/A'
                ))
                changed = True
            elif code == 'PRIMEROS_AUXILIOS':
                permit = existing_types[code]
                if permit.applicability == 'N/A' and permit.expiration_date is None:
                    permit.applicability = 'YES'
                    changed = True
        if changed:
            db.session.commit()

    all_permits = emp.permits.order_by(EmployeePermit.permit_type).all()
    active_permits = [p for p in all_permits if p.applicability != 'N/A']
    hidden_permits = [p for p in all_permits if p.applicability == 'N/A']
    return render_template('employee.html', employee=emp,
                           permits=active_permits, hidden_permits=hidden_permits,
                           permit_types=EMPLOYEE_PERMIT_TYPES,
                           nav=_sibling_nav('employees', emp.id, request.args))


@app.route('/employee/new', methods=['GET', 'POST'])
@manager_required
def employee_new():
    if request.method == 'POST':
        emp = Employee(
            name=request.form['name'],
            company=request.form['company'],
            area=request.form.get('area', ''),
            status=request.form.get('status', 'activo'),
            license_number=request.form.get('license_number', ''),
            puesto=request.form.get('puesto', ''),
            telefono=request.form.get('telefono', ''),
            email=request.form.get('email', ''),
            contacto_emergencia=request.form.get('contacto_emergencia', ''),
            shirt_size=request.form.get('shirt_size', ''),
            endoso_hazmat=request.form.get('endoso_hazmat', 'N/A'),
        )
        # Parse dates
        for field in ['fecha_nacimiento', 'license_expiration', 'fecha_contratacion']:
            val = request.form.get(field, '')
            if val:
                try:
                    setattr(emp, field, datetime.strptime(val, '%Y-%m-%d').date())
                except ValueError:
                    pass

        # Handle license file upload
        if 'license_file' in request.files:
            file = request.files['license_file']
            if file and file.filename and allowed_file(file.filename):
                if emp.license_file:
                    delete_stored_file(emp.license_file)
                filename = save_uploaded_file(file)
                emp.license_file = filename

        db.session.add(emp)
        db.session.flush()  # Get ID

        # Create default permit slots
        for code, name in EMPLOYEE_PERMIT_TYPES:
            if code != 'OTHER':
                permit = EmployeePermit(
                    employee_id=emp.id,
                    permit_type=code,
                    applicability='YES' if code == 'PRIMEROS_AUXILIOS' else 'N/A'
                )
                db.session.add(permit)

        db.session.commit()
        flash(f'Empleado {emp.name} creado.', 'success')
        return redirect(url_for('employee_detail', id=emp.id))

    return render_template('employee_form.html', employee=None)


@app.route('/employee/<int:id>/edit', methods=['GET', 'POST'])
@manager_required
def employee_edit(id):
    emp = Employee.query.get_or_404(id)
    if request.method == 'POST':
        emp.name = request.form['name']
        emp.company = request.form['company']
        emp.area = request.form.get('area', '')
        if not emp.is_archived:
            emp.status = request.form.get('status', 'activo')
        emp.license_number = request.form.get('license_number', '')
        emp.puesto = request.form.get('puesto', '')
        emp.telefono = request.form.get('telefono', '')
        emp.email = request.form.get('email', '')
        emp.contacto_emergencia = request.form.get('contacto_emergencia', '')
        emp.shirt_size = request.form.get('shirt_size', '')
        emp.endoso_hazmat = request.form.get('endoso_hazmat', 'N/A')

        for field in ['fecha_nacimiento', 'license_expiration', 'fecha_contratacion']:
            val = request.form.get(field, '')
            if val:
                try:
                    setattr(emp, field, datetime.strptime(val, '%Y-%m-%d').date())
                except ValueError:
                    pass
            else:
                setattr(emp, field, None)

        if 'license_file' in request.files:
            file = request.files['license_file']
            if file and file.filename and allowed_file(file.filename):
                if emp.license_file:
                    delete_stored_file(emp.license_file)
                filename = save_uploaded_file(file)
                emp.license_file = filename

        db.session.commit()
        flash(f'Empleado {emp.name} actualizado.', 'success')
        return redirect(url_for('employee_detail', id=emp.id))

    return render_template('employee_form.html', employee=emp)


@app.route('/employee/<int:id>/delete', methods=['POST'])
@admin_required
def employee_delete(id):
    emp = Employee.query.get_or_404(id)
    name = emp.name
    # Explicitly delete child permits first: the relationship is lazy='dynamic'
    # and the FK has no ON DELETE CASCADE, so the ORM-level cascade doesn't
    # always issue child DELETEs before the parent, causing FK violations.
    EmployeePermit.query.filter_by(employee_id=emp.id).delete(synchronize_session=False)
    db.session.delete(emp)
    db.session.commit()
    flash(f'Empleado {name} eliminado.', 'warning')
    return redirect(url_for('dashboard'))


@app.route('/employee/<int:id>/archive', methods=['POST'])
@manager_required
def employee_archive(id):
    emp = Employee.query.get_or_404(id)
    if emp.is_archived:
        flash('Este empleado ya está archivado.', 'warning')
        return redirect(url_for('employee_detail', id=emp.id))
    reason = request.form.get('archive_reason', 'otro')
    if reason not in {code for code, _ in EMPLOYEE_ARCHIVE_REASONS}:
        reason = 'otro'
    emp.archive_reason = reason
    emp.archive_note = (request.form.get('archive_note') or '').strip()[:300] or None
    emp.archived_at = datetime.utcnow()
    emp.status = 'inactivo'
    db.session.commit()
    flash(f'Empleado {emp.name} archivado.', 'warning')
    return redirect(url_for('employee_detail', id=emp.id))


@app.route('/employee/<int:id>/restore', methods=['POST'])
@manager_required
def employee_restore(id):
    emp = Employee.query.get_or_404(id)
    if not emp.is_archived:
        flash('Este empleado no está archivado.', 'warning')
        return redirect(url_for('employee_detail', id=emp.id))
    emp.archived_at = None
    emp.archive_reason = None
    emp.archive_note = None
    emp.status = 'activo'
    db.session.commit()
    flash(f'Empleado {emp.name} restaurado.', 'success')
    return redirect(url_for('employee_detail', id=emp.id))


# ── EMPLOYEE PERMIT ROUTES ─────────────────────────────────────────────

@app.route('/employee/<int:emp_id>/permit/<int:permit_id>/edit', methods=['POST'])
@manager_required
def employee_permit_edit(emp_id, permit_id):
    permit = EmployeePermit.query.get_or_404(permit_id)
    if permit.employee_id != emp_id:
        abort(403)

    permit.applicability = request.form.get('applicability', 'YES')
    permit.permit_number = request.form.get('permit_number', '')
    permit.issuing_authority = request.form.get('issuing_authority', '')
    permit.notes = request.form.get('notes', '')

    val = request.form.get('expiration_date', '')
    if val:
        try:
            permit.expiration_date = datetime.strptime(val, '%Y-%m-%d').date()
        except ValueError:
            pass
    else:
        permit.expiration_date = None

    cost = request.form.get('renewal_cost', '')
    if cost:
        try:
            permit.renewal_cost = float(cost)
        except ValueError:
            pass

    if 'permit_file' in request.files:
        file = request.files['permit_file']
        if file and file.filename and allowed_file(file.filename):
            if permit.file_path:
                delete_stored_file(permit.file_path)
            filename = save_uploaded_file(file)
            permit.file_path = filename

    db.session.commit()
    flash(f'Permiso {permit.display_name} actualizado.', 'success')
    return redirect(url_for('employee_detail', id=emp_id))


@app.route('/employee/<int:emp_id>/permit/<int:permit_id>/toggle', methods=['POST'])
@manager_required
def employee_permit_toggle(emp_id, permit_id):
    permit = EmployeePermit.query.get_or_404(permit_id)
    if permit.employee_id != emp_id:
        abort(403)
    new_value = request.form.get('applicability', 'YES')
    permit.applicability = new_value
    if new_value == 'N/A':
        permit.expiration_date = None
    db.session.commit()
    flash(f'Permiso {permit.display_name} {"activado" if new_value == "YES" else "desactivado"}.', 'success')
    return redirect(url_for('employee_detail', id=emp_id))


@app.route('/employee/<int:emp_id>/permit/<int:permit_id>/upload-form', methods=['POST'])
@manager_required
def employee_permit_upload_form(emp_id, permit_id):
    permit = EmployeePermit.query.get_or_404(permit_id)
    if permit.employee_id != emp_id:
        abort(403)

    file = request.files.get('form_file')
    if not file or not file.filename:
        flash('No se seleccionó ningún archivo.', 'error')
        return redirect(url_for('employee_detail', id=emp_id))
    if '.' not in file.filename or file.filename.rsplit('.', 1)[1].lower() != 'pdf':
        flash('Solo se permiten archivos PDF.', 'error')
        return redirect(url_for('employee_detail', id=emp_id))

    if permit.file_path:
        delete_stored_file(permit.file_path)

    filename = save_uploaded_file(file)
    permit.file_path = filename
    db.session.commit()
    flash(f'Formulario adjuntado a {permit.display_name}.', 'success')
    return redirect(url_for('employee_detail', id=emp_id))


@app.route('/employee/<int:emp_id>/permit/<int:permit_id>/delete-form', methods=['POST'])
@manager_required
def employee_permit_delete_form(emp_id, permit_id):
    permit = EmployeePermit.query.get_or_404(permit_id)
    if permit.employee_id != emp_id:
        abort(403)

    if permit.file_path:
        delete_stored_file(permit.file_path)
        permit.file_path = None
        db.session.commit()
        flash(f'Formulario eliminado de {permit.display_name}.', 'success')
    return redirect(url_for('employee_detail', id=emp_id))


# ── LICENSE DOCUMENT ROUTES ────────────────────────────────────────────

@app.route('/employee/<int:id>/license/upload-form', methods=['POST'])
@manager_required
def employee_license_upload_form(id):
    emp = Employee.query.get_or_404(id)

    file = request.files.get('form_file')
    if not file or not file.filename:
        flash('No se seleccionó ningún archivo.', 'error')
        return redirect(url_for('employee_detail', id=emp.id))
    if not allowed_file(file.filename):
        flash('Tipo de archivo no permitido.', 'error')
        return redirect(url_for('employee_detail', id=emp.id))

    if emp.license_file:
        delete_stored_file(emp.license_file)

    emp.license_file = save_uploaded_file(file)
    db.session.commit()
    flash('Documento de licencia adjuntado.', 'success')
    return redirect(url_for('employee_detail', id=emp.id))


@app.route('/employee/<int:id>/license/delete-form', methods=['POST'])
@manager_required
def employee_license_delete_form(id):
    emp = Employee.query.get_or_404(id)

    if emp.license_file:
        delete_stored_file(emp.license_file)
        emp.license_file = None
        db.session.commit()
        flash('Documento de licencia eliminado.', 'success')
    return redirect(url_for('employee_detail', id=emp.id))


@app.route('/employee/<int:emp_id>/permit/new', methods=['POST'])
@manager_required
def employee_permit_new(emp_id):
    emp = Employee.query.get_or_404(emp_id)
    permit = EmployeePermit(
        employee_id=emp.id,
        permit_type=request.form.get('permit_type', 'OTHER'),
        permit_name=request.form.get('permit_name', ''),
        applicability='YES',
    )
    val = request.form.get('expiration_date', '')
    if val:
        try:
            permit.expiration_date = datetime.strptime(val, '%Y-%m-%d').date()
        except ValueError:
            pass

    db.session.add(permit)
    db.session.commit()
    flash(f'Nuevo permiso agregado.', 'success')
    return redirect(url_for('employee_detail', id=emp_id))


# ── EQUIPMENT ROUTES ───────────────────────────────────────────────────

@app.route('/equipment/<int:id>')
@login_required
def equipment_detail(id):
    equip = Equipment.query.get_or_404(id)
    all_permits = equip.permits.order_by(EquipmentPermit.permit_type).all()
    active_permits = [p for p in all_permits if p.applicability != 'N/A']
    hidden_permits = [p for p in all_permits if p.applicability == 'N/A']

    # Insurance is normally shared at the company level for LB/PLI — show the
    # policy read-only and drop the redundant per-vehicle SEGURO permit. A vehicle
    # set to its own policy resolves to None and keeps its editable SEGURO card.
    shared_insurance = None
    shared_insurance_title = None
    insurance_type = equip.insurance_permit_type
    if insurance_type:
        shared_insurance = CompanyPermit.query.filter_by(
            company=equip.company, permit_type=insurance_type).first()
        shared_insurance_title = equip.insurance_card_title
        active_permits = [p for p in active_permits if p.permit_type != 'SEGURO']
        hidden_permits = [p for p in hidden_permits if p.permit_type != 'SEGURO']

    # Full vehicle history: this equipment's issue reports (newest first).
    issues = equip.issues.order_by(Issue.reported_at.desc()).all()

    return render_template('equipment.html', equipment=equip,
                           permits=active_permits, hidden_permits=hidden_permits,
                           shared_insurance=shared_insurance,
                           shared_insurance_title=shared_insurance_title,
                           permit_types=EQUIPMENT_PERMIT_TYPES,
                           issues=issues,
                           nav=_sibling_nav('equipment', equip.id, request.args))


# ── EQUIPMENT SHEET COLUMN MAPPING ─────────────────────────────────────
# The equipment import locates every column by its header text instead of a
# fixed position. Aliases are matched against accent-stripped, upper-cased
# headers: exact match first, then substring, so "SEGURO" can't swallow a
# more specific column. Add new spellings here, not new positional indexes.
EQUIPMENT_COLUMN_ALIASES = [
    ('titular',    ('TITULAR', 'TITULO', 'DUENO', 'PROPIETARIO')),
    ('unit',       ('UNIDAD', 'NO UNIDAD', 'NUM UNIDAD', 'UNIT')),
    ('company',    ('COMPANIA', 'COMPANIA DUENA', 'EMPRESA', 'COMPANY', 'CIA')),
    ('clase',      ('CLASE', 'TIPO')),
    ('make',       ('MARCA', 'MAKE')),
    ('model',      ('MODELO', 'MODEL')),
    ('year',       ('ANO', 'YEAR')),
    ('vin',        ('VIN', 'SERIAL', 'NO SERIE')),
    ('plate',      ('TABLILLA', 'PLACA', 'PLATE')),
    ('insurance',  ('SEGURO', 'ASEGURADORA', 'INSURANCE')),
    ('MARBETE',    ('MARBETE',)),
    ('INSPECCION', ('INSPECCION', 'INSPECION')),
    ('NTSP',       ('NTSP',)),
    ('cost',       ('COSTO', 'PRECIO', 'COST')),
]

# Human labels for the abort message when a required column can't be found.
EQUIPMENT_REQUIRED_COLUMNS = [('company', 'Compañía')]
# At least one identifying column must be present, otherwise every row would
# look like a brand-new vehicle.
EQUIPMENT_IDENTITY_COLUMNS = [('unit', 'Unidad'), ('vin', 'VIN'), ('plate', 'Tablilla')]


def _normalize_header(value):
    """Accent-stripped, upper-cased, single-spaced header text."""
    if value is None:
        return ''
    s = unicodedata.normalize('NFKD', str(value)).encode('ascii', 'ignore').decode()
    s = re.sub(r'[^A-Za-z0-9]+', ' ', s)
    return ' '.join(s.split()).upper()


def _map_equipment_columns(normalized_headers):
    """Map canonical field names to column indexes by header text."""
    mapping = {}
    used = set()
    for exact in (True, False):
        for field, aliases in EQUIPMENT_COLUMN_ALIASES:
            if field in mapping:
                continue
            for idx, header in enumerate(normalized_headers):
                if idx in used or not header:
                    continue
                hit = (header in aliases) if exact else any(a in header for a in aliases)
                if hit:
                    mapping[field] = idx
                    used.add(idx)
                    break
    return mapping


def _missing_equipment_columns(columns):
    """Spanish labels of the columns the sheet must have but doesn't."""
    missing = [label for field, label in EQUIPMENT_REQUIRED_COLUMNS if field not in columns]
    if not any(field in columns for field, _ in EQUIPMENT_IDENTITY_COLUMNS):
        missing.append(' / '.join(label for _, label in EQUIPMENT_IDENTITY_COLUMNS))
    return missing


def _resolve_company(*cells):
    """First recognizable company among the given cells, else None.

    Returning None (instead of defaulting to LB) is deliberate: a blank or
    unreadable Compañía cell must stop the row, not misfile the vehicle.
    """
    for cell in cells:
        raw = _normalize_header(cell)
        if not raw:
            continue
        if 'PERSONAL' in raw:
            return 'Personal'
        if 'PLI' in raw or 'PROFESSIONAL' in raw or 'LOGISTIC' in raw:
            return 'PLI'
        if raw == 'LB' or raw.startswith('LB ') or 'CARIBE' in raw:
            return 'LB'
    return None


def _resolve_import_class(class_cell, model):
    """Equipment class from the sheet's Clase/Tipo column, else from the model."""
    raw = _normalize_header(class_cell)
    if raw:
        for code, label in EQUIPMENT_CLASSES:
            if raw == code.upper() or raw == _normalize_header(label):
                return code
        inferred = classify_equipment(raw)
        if inferred != 'truck' or raw == _normalize_header('Camión'):
            return inferred
    return classify_equipment(model)


def find_duplicate_equipment(vin_serial, plate_number, unit_number, company,
                             titular=None, model=None, year=None):
    """Return an existing Equipment matching by VIN, then plate, then unit#+company.

    When a vehicle has no VIN, plate, or unit number (e.g. carretones, generadores,
    tanques), fall back to a normalized (company, titular, model, year) match so
    re-imports of these keyless rows don't create complete duplicates.
    """
    vin = (vin_serial or '').strip()
    if vin:
        m = Equipment.query.filter(db.func.lower(Equipment.vin_serial) == vin.lower()).first()
        if m:
            return m
    plate = (plate_number or '').strip()
    if plate:
        m = Equipment.query.filter(db.func.lower(Equipment.plate_number) == plate.lower()).first()
        if m:
            return m
    unit = (unit_number or '').strip()
    if unit:
        return Equipment.query.filter(
            db.func.lower(Equipment.unit_number) == unit.lower(),
            Equipment.company == company,
        ).first()

    # Keyless fallback: no VIN/plate/unit. Only attempt when there is enough to
    # identify the row (titular or model present) so we never match on an empty tuple.
    titular_n = (titular or '').strip()
    model_n = (model or '').strip()
    if titular_n or model_n:
        return Equipment.query.filter(
            (Equipment.vin_serial.is_(None) | (db.func.trim(Equipment.vin_serial) == '')),
            (Equipment.plate_number.is_(None) | (db.func.trim(Equipment.plate_number) == '')),
            (Equipment.unit_number.is_(None) | (db.func.trim(Equipment.unit_number) == '')),
            Equipment.company == company,
            db.func.lower(db.func.coalesce(Equipment.titular, '')) == titular_n.lower(),
            db.func.lower(db.func.coalesce(Equipment.model, '')) == model_n.lower(),
            (Equipment.year == year) if year is not None else Equipment.year.is_(None),
        ).first()
    return None


def _resolve_equipment_class(form):
    """Validated class from the form select, falling back to the model-based classifier."""
    cls = form.get('equipment_class', '')
    if cls not in {code for code, _ in EQUIPMENT_CLASSES}:
        cls = classify_equipment(form.get('model', ''))
    return cls


def _resolve_insurance_policy(form, company):
    """Validated insurance choice from the form select.

    None = automática (follow the equipment class); a SEGURO_* code = that shared
    company policy; INSURANCE_OWN = the vehicle's own per-vehicle SEGURO permit.
    Unknown values silently reset to automática (the dashboard filter idiom).
    Personal equipment has no shared policies, so it is always on its own.
    """
    if company not in ('LB', 'PLI'):
        return INSURANCE_OWN
    choice = form.get('insurance_policy_type', '')
    if choice == INSURANCE_OWN or choice in set(INSURANCE_TYPE_BY_CLASS.values()):
        return choice
    return None


@app.route('/equipment/new', methods=['GET', 'POST'])
@manager_required
def equipment_new():
    if request.method == 'POST':
        form_year = None
        if request.form.get('year', '').strip():
            try:
                form_year = int(request.form['year'])
            except ValueError:
                form_year = None
        match = find_duplicate_equipment(
            request.form.get('vin_serial', ''),
            request.form.get('plate_number', ''),
            request.form.get('unit_number', ''),
            request.form.get('company', ''),
            titular=request.form.get('titular', ''),
            model=request.form.get('model', ''),
            year=form_year,
        )
        if match:
            archived_hint = (' (equipo archivado — restáurelo o elimínelo antes de '
                             'reutilizar sus datos)') if match.is_archived else ''
            flash(
                f'Ya existe un equipo con esa unidad/placa/VIN: {match.display_name} '
                f'({match.company_full}){archived_hint}. Verifique antes de crear un duplicado.',
                'danger',
            )
            return render_template('equipment_form.html', equipment=None,
                                   form_data=request.form, duplicate=match,
                                   equipment_classes=EQUIPMENT_CLASSES,
                                   insurance_policy_choices=INSURANCE_POLICY_CHOICES), 422
        equip = Equipment(
            company=request.form['company'],
            equipment_class=_resolve_equipment_class(request.form),
            insurance_policy_type=_resolve_insurance_policy(
                request.form, request.form['company']),
            titular=request.form.get('titular', ''),
            name=request.form.get('name', ''),
            unit_number=request.form.get('unit_number', ''),
            plate_number=request.form.get('plate_number', ''),
            make=request.form.get('make', ''),
            model=request.form.get('model', ''),
            vin_serial=request.form.get('vin_serial', ''),
            insurance_company=request.form.get('insurance_company', ''),
            notes=request.form.get('notes', ''),
            status=request.form.get('status', 'activo'),
        )
        yr = request.form.get('year', '')
        if yr:
            try:
                equip.year = int(yr)
            except ValueError:
                pass

        cost_raw = request.form.get('cost', '').strip()
        if cost_raw:
            try:
                equip.cost = Decimal(cost_raw)
            except (InvalidOperation, ValueError):
                pass

        db.session.add(equip)
        db.session.flush()

        # Create default permit slots
        for code, name in EQUIPMENT_PERMIT_TYPES:
            if code != 'OTHER':
                permit = EquipmentPermit(
                    equipment_id=equip.id,
                    permit_type=code,
                    applicability='YES',
                )
                db.session.add(permit)

        # SEGURO goes N/A when a shared company policy covers the vehicle.
        sync_vehicle_insurance_permit(equip)

        db.session.commit()
        flash(f'Equipo {equip.display_name} creado.', 'success')
        return redirect(url_for('equipment_detail', id=equip.id))

    return render_template('equipment_form.html', equipment=None,
                           equipment_classes=EQUIPMENT_CLASSES,
                           insurance_policy_choices=INSURANCE_POLICY_CHOICES)


@app.route('/equipment/<int:id>/edit', methods=['GET', 'POST'])
@manager_required
def equipment_edit(id):
    equip = Equipment.query.get_or_404(id)
    if request.method == 'POST':
        equip.company = request.form['company']
        equip.equipment_class = _resolve_equipment_class(request.form)
        equip.insurance_policy_type = _resolve_insurance_policy(
            request.form, equip.company)
        equip.titular = request.form.get('titular', '')
        equip.name = request.form.get('name', '')
        equip.unit_number = request.form.get('unit_number', '')
        equip.plate_number = request.form.get('plate_number', '')
        equip.make = request.form.get('make', '')
        equip.model = request.form.get('model', '')
        equip.vin_serial = request.form.get('vin_serial', '')
        equip.insurance_company = request.form.get('insurance_company', '')
        equip.notes = request.form.get('notes', '')
        if not equip.is_archived:
            equip.status = request.form.get('status', 'activo')

        yr = request.form.get('year', '')
        if yr:
            try:
                equip.year = int(yr)
            except ValueError:
                pass

        cost_raw = request.form.get('cost', '').strip()
        if cost_raw:
            try:
                equip.cost = Decimal(cost_raw)
            except (InvalidOperation, ValueError):
                pass
        else:
            equip.cost = None

        # Company/class/policy changes can flip which side owns the insurance.
        sync_vehicle_insurance_permit(equip)

        db.session.commit()
        flash(f'Equipo {equip.display_name} actualizado.', 'success')
        return redirect(url_for('equipment_detail', id=equip.id))

    return render_template('equipment_form.html', equipment=equip,
                           equipment_classes=EQUIPMENT_CLASSES,
                           insurance_policy_choices=INSURANCE_POLICY_CHOICES)


@app.route('/equipment/<int:id>/delete', methods=['POST'])
@admin_required
def equipment_delete(id):
    equip = Equipment.query.get_or_404(id)
    name = equip.display_name
    # Explicitly delete child permits first (see employee_delete for rationale).
    EquipmentPermit.query.filter_by(equipment_id=equip.id).delete(synchronize_session=False)
    db.session.delete(equip)
    db.session.commit()
    flash(f'Equipo {name} eliminado.', 'warning')
    return redirect(url_for('dashboard', view='equipment'))


@app.route('/equipment/<int:id>/archive', methods=['POST'])
@manager_required
def equipment_archive(id):
    equip = Equipment.query.get_or_404(id)
    if equip.is_archived:
        flash('Este equipo ya está archivado.', 'warning')
        return redirect(url_for('equipment_detail', id=equip.id))
    reason = request.form.get('archive_reason', 'otro')
    if reason not in {code for code, _ in EQUIPMENT_ARCHIVE_REASONS}:
        reason = 'otro'
    equip.archive_reason = reason
    equip.archive_note = (request.form.get('archive_note') or '').strip()[:300] or None
    equip.archived_at = datetime.utcnow()
    equip.status = 'inactivo'
    db.session.commit()
    flash(f'Equipo {equip.display_name} archivado.', 'warning')
    return redirect(url_for('equipment_detail', id=equip.id))


@app.route('/equipment/<int:id>/restore', methods=['POST'])
@manager_required
def equipment_restore(id):
    equip = Equipment.query.get_or_404(id)
    if not equip.is_archived:
        flash('Este equipo no está archivado.', 'warning')
        return redirect(url_for('equipment_detail', id=equip.id))
    equip.archived_at = None
    equip.archive_reason = None
    equip.archive_note = None
    equip.status = 'activo'
    db.session.commit()
    flash(f'Equipo {equip.display_name} restaurado.', 'success')
    return redirect(url_for('equipment_detail', id=equip.id))


@app.route('/equipment/<int:eq_id>/permit/<int:permit_id>/edit', methods=['POST'])
@manager_required
def equipment_permit_edit(eq_id, permit_id):
    permit = EquipmentPermit.query.get_or_404(permit_id)
    if permit.equipment_id != eq_id:
        abort(403)

    permit.applicability = request.form.get('applicability', 'YES')
    permit.permit_number = request.form.get('permit_number', '')
    permit.issuing_authority = request.form.get('issuing_authority', '')
    permit.notes = request.form.get('notes', '')

    val = request.form.get('expiration_date', '')
    if val:
        try:
            permit.expiration_date = datetime.strptime(val, '%Y-%m-%d').date()
        except ValueError:
            pass
    else:
        permit.expiration_date = None

    cost = request.form.get('renewal_cost', '')
    if cost:
        try:
            permit.renewal_cost = float(cost)
        except ValueError:
            pass

    if 'permit_file' in request.files:
        file = request.files['permit_file']
        if file and file.filename and allowed_file(file.filename):
            if permit.file_path:
                delete_stored_file(permit.file_path)
            filename = save_uploaded_file(file)
            permit.file_path = filename

    db.session.commit()
    flash(f'Permiso {permit.display_name} actualizado.', 'success')
    return redirect(url_for('equipment_detail', id=eq_id))


@app.route('/equipment/<int:eq_id>/permit/<int:permit_id>/toggle', methods=['POST'])
@manager_required
def equipment_permit_toggle(eq_id, permit_id):
    permit = EquipmentPermit.query.get_or_404(permit_id)
    if permit.equipment_id != eq_id:
        abort(403)
    new_value = request.form.get('applicability', 'YES')
    permit.applicability = new_value
    if new_value == 'N/A':
        permit.expiration_date = None
    db.session.commit()
    flash(f'Permiso {permit.display_name} {"activado" if new_value == "YES" else "desactivado"}.', 'success')
    return redirect(url_for('equipment_detail', id=eq_id))


@app.route('/equipment/<int:eq_id>/permit/<int:permit_id>/upload-form', methods=['POST'])
@manager_required
def equipment_permit_upload_form(eq_id, permit_id):
    permit = EquipmentPermit.query.get_or_404(permit_id)
    if permit.equipment_id != eq_id:
        abort(403)

    file = request.files.get('form_file')
    if not file or not file.filename:
        flash('No se seleccionó ningún archivo.', 'error')
        return redirect(url_for('equipment_detail', id=eq_id))
    if '.' not in file.filename or file.filename.rsplit('.', 1)[1].lower() != 'pdf':
        flash('Solo se permiten archivos PDF.', 'error')
        return redirect(url_for('equipment_detail', id=eq_id))

    if permit.file_path:
        delete_stored_file(permit.file_path)

    filename = save_uploaded_file(file)
    permit.file_path = filename
    db.session.commit()
    flash(f'Formulario adjuntado a {permit.display_name}.', 'success')
    return redirect(url_for('equipment_detail', id=eq_id))


@app.route('/equipment/<int:eq_id>/permit/<int:permit_id>/delete-form', methods=['POST'])
@manager_required
def equipment_permit_delete_form(eq_id, permit_id):
    permit = EquipmentPermit.query.get_or_404(permit_id)
    if permit.equipment_id != eq_id:
        abort(403)

    if permit.file_path:
        delete_stored_file(permit.file_path)
        permit.file_path = None
        db.session.commit()
        flash(f'Formulario eliminado de {permit.display_name}.', 'success')
    return redirect(url_for('equipment_detail', id=eq_id))


@app.route('/equipment/<int:eq_id>/permit/new', methods=['POST'])
@manager_required
def equipment_permit_new(eq_id):
    equip = Equipment.query.get_or_404(eq_id)
    # A stale open tab could post a retired type — anything not in the current
    # vocabulary falls back to OTHER.
    permit_type = request.form.get('permit_type', 'OTHER')
    if permit_type not in {code for code, _ in EQUIPMENT_PERMIT_TYPES}:
        permit_type = 'OTHER'
    permit = EquipmentPermit(
        equipment_id=equip.id,
        permit_type=permit_type,
        permit_name=request.form.get('permit_name', ''),
        applicability='YES',
    )
    val = request.form.get('expiration_date', '')
    if val:
        try:
            permit.expiration_date = datetime.strptime(val, '%Y-%m-%d').date()
        except ValueError:
            pass
    db.session.add(permit)
    db.session.commit()
    flash(f'Nuevo permiso agregado.', 'success')
    return redirect(url_for('equipment_detail', id=eq_id))


# ── COMPANY PERMIT ROUTES ─────────────────────────────────────────────

@app.route('/company/<company>/permit/<int:permit_id>/edit', methods=['POST'])
@manager_required
def company_permit_edit(company, permit_id):
    permit = CompanyPermit.query.get_or_404(permit_id)
    if permit.company != company:
        abort(403)

    permit.applicability = request.form.get('applicability', 'YES')
    permit.permit_number = request.form.get('permit_number', '')
    permit.issuing_authority = request.form.get('issuing_authority', '')
    permit.notes = request.form.get('notes', '')

    val = request.form.get('expiration_date', '')
    if val:
        try:
            permit.expiration_date = datetime.strptime(val, '%Y-%m-%d').date()
        except ValueError:
            pass
    else:
        permit.expiration_date = None

    cost_val = request.form.get('renewal_cost', '')
    if cost_val:
        try:
            permit.renewal_cost = float(cost_val)
        except (ValueError, TypeError):
            pass
    else:
        permit.renewal_cost = None

    file = request.files.get('permit_file')
    if file and file.filename:
        if permit.file_path:
            delete_stored_file(permit.file_path)
        permit.file_path = save_uploaded_file(file)

    db.session.commit()
    flash(f'{permit.display_name} actualizado.', 'success')
    return redirect(url_for('dashboard', view='company'))


@app.route('/company/<company>/permit/<int:permit_id>/upload-form', methods=['POST'])
@manager_required
def company_permit_upload_form(company, permit_id):
    permit = CompanyPermit.query.get_or_404(permit_id)
    if permit.company != company:
        abort(403)

    file = request.files.get('form_file')
    if not file or not file.filename:
        flash('No se seleccionó archivo.', 'warning')
        return redirect(url_for('dashboard', view='company'))

    if not file.filename.lower().endswith('.pdf'):
        flash('Solo se permiten archivos PDF.', 'warning')
        return redirect(url_for('dashboard', view='company'))

    if permit.file_path:
        delete_stored_file(permit.file_path)

    filename = save_uploaded_file(file)
    permit.file_path = filename
    db.session.commit()
    flash(f'Formulario adjuntado a {permit.display_name}.', 'success')
    return redirect(url_for('dashboard', view='company'))


@app.route('/company/<company>/permit/<int:permit_id>/delete-form', methods=['POST'])
@manager_required
def company_permit_delete_form(company, permit_id):
    permit = CompanyPermit.query.get_or_404(permit_id)
    if permit.company != company:
        abort(403)

    if permit.file_path:
        delete_stored_file(permit.file_path)
        permit.file_path = None
        db.session.commit()
        flash(f'Formulario eliminado de {permit.display_name}.', 'success')
    return redirect(url_for('dashboard', view='company'))


# ── FILE SERVING ───────────────────────────────────────────────────────

@app.route('/uploads/<filename>')
@login_required
def uploaded_file(filename):
    stored = FileStorage.query.filter_by(filename=filename).first_or_404()
    return send_file(
        BytesIO(stored.data),
        mimetype=stored.mime_type,
        as_attachment=False,
        download_name=stored.original_filename or filename,
    )


# ── DEBUG: Duplicate Check & Fix (temporary) ──────────────────────────

@app.route('/debug/duplicates')
@login_required
def debug_duplicates():
    dupes = db.session.query(
        Employee.name, Employee.company, db.func.count(Employee.id)
    ).group_by(Employee.name, Employee.company).having(db.func.count(Employee.id) > 1).all()
    return {'duplicates': [{'name': d[0], 'company': d[1], 'count': d[2]} for d in dupes],
            'total_duplicate_groups': len(dupes)}


@app.route('/debug/dedup', methods=['POST'])
@admin_required
def run_dedup():
    """Web-accessible dedup for environments without CLI access."""
    emp_deleted = 0
    permit_deleted = 0
    eq_permit_deleted = 0

    # --- Deduplicate employees by (name, company) ---
    dupes = (
        db.session.query(Employee.name, Employee.company, db.func.count(Employee.id))
        .group_by(Employee.name, Employee.company)
        .having(db.func.count(Employee.id) > 1)
        .all()
    )
    for name, company, count in dupes:
        employees = Employee.query.filter_by(name=name, company=company).order_by(Employee.id).all()
        keeper = employees[0]
        for dup in employees[1:]:
            for col in ['area', 'status', 'fecha_nacimiento', 'license_number',
                        'license_expiration', 'license_file', 'puesto', 'telefono',
                        'email', 'fecha_contratacion', 'contacto_emergencia', 'shirt_size',
                        'archived_at', 'archive_reason', 'archive_note']:
                if getattr(keeper, col) is None and getattr(dup, col) is not None:
                    setattr(keeper, col, getattr(dup, col))
            if keeper.endoso_hazmat == 'N/A' and dup.endoso_hazmat != 'N/A':
                keeper.endoso_hazmat = dup.endoso_hazmat
            for dup_permit in dup.permits.all():
                keeper_permit = keeper.permits.filter_by(permit_type=dup_permit.permit_type).first()
                if keeper_permit:
                    if keeper_permit.expiration_date is None and dup_permit.expiration_date is not None:
                        keeper_permit.expiration_date = dup_permit.expiration_date
                    if keeper_permit.applicability == 'N/A' and dup_permit.applicability == 'YES':
                        keeper_permit.applicability = dup_permit.applicability
                    for field in ['permit_number', 'issuing_authority', 'file_path', 'renewal_cost', 'notes']:
                        if getattr(keeper_permit, field) is None and getattr(dup_permit, field) is not None:
                            setattr(keeper_permit, field, getattr(dup_permit, field))
            db.session.delete(dup)
            emp_deleted += 1

    # --- Deduplicate employee permits ---
    permit_dupes = (
        db.session.query(EmployeePermit.employee_id, EmployeePermit.permit_type, db.func.count(EmployeePermit.id))
        .filter(EmployeePermit.permit_type != 'OTHER')
        .group_by(EmployeePermit.employee_id, EmployeePermit.permit_type)
        .having(db.func.count(EmployeePermit.id) > 1)
        .all()
    )
    for emp_id, ptype, count in permit_dupes:
        permits = EmployeePermit.query.filter_by(employee_id=emp_id, permit_type=ptype).order_by(EmployeePermit.id).all()
        keeper = permits[0]
        for dup in permits[1:]:
            if keeper.expiration_date is None and dup.expiration_date is not None:
                keeper.expiration_date = dup.expiration_date
            if keeper.applicability == 'N/A' and dup.applicability == 'YES':
                keeper.applicability = dup.applicability
            for field in ['permit_number', 'issuing_authority', 'file_path', 'renewal_cost', 'notes']:
                if getattr(keeper, field) is None and getattr(dup, field) is not None:
                    setattr(keeper, field, getattr(dup, field))
            db.session.delete(dup)
            permit_deleted += 1

    # --- Deduplicate equipment permits ---
    eq_permit_dupes = (
        db.session.query(EquipmentPermit.equipment_id, EquipmentPermit.permit_type, db.func.count(EquipmentPermit.id))
        .filter(EquipmentPermit.permit_type != 'OTHER')
        .group_by(EquipmentPermit.equipment_id, EquipmentPermit.permit_type)
        .having(db.func.count(EquipmentPermit.id) > 1)
        .all()
    )
    for eq_id, ptype, count in eq_permit_dupes:
        permits = EquipmentPermit.query.filter_by(equipment_id=eq_id, permit_type=ptype).order_by(EquipmentPermit.id).all()
        keeper = permits[0]
        for dup in permits[1:]:
            if keeper.expiration_date is None and dup.expiration_date is not None:
                keeper.expiration_date = dup.expiration_date
            if keeper.applicability == 'N/A' and dup.applicability == 'YES':
                keeper.applicability = dup.applicability
            for field in ['permit_number', 'issuing_authority', 'file_path', 'renewal_cost', 'notes']:
                if getattr(keeper, field) is None and getattr(dup, field) is not None:
                    setattr(keeper, field, getattr(dup, field))
            db.session.delete(dup)
            eq_permit_deleted += 1

    db.session.commit()

    # Apply unique constraints
    if db.engine.dialect.name == 'postgresql':
        for sql in [
            "ALTER TABLE employees ADD CONSTRAINT uq_employee_name_company UNIQUE (name, company)",
            "ALTER TABLE employee_permits ADD CONSTRAINT uq_employee_permit_type UNIQUE (employee_id, permit_type)",
            "ALTER TABLE equipment_permits ADD CONSTRAINT uq_equipment_permit_type UNIQUE (equipment_id, permit_type)",
        ]:
            try:
                db.session.execute(db.text(sql))
                db.session.commit()
            except Exception:
                db.session.rollback()

    return {
        'status': 'ok',
        'employees_deleted': emp_deleted,
        'employee_permits_deleted': permit_deleted,
        'equipment_permits_deleted': eq_permit_deleted,
    }


# ── EXCEL IMPORT ───────────────────────────────────────────────────────

@app.route('/import', methods=['GET', 'POST'])
@admin_required
@protect_uploaded_files()   # documents are manual-only; see models.py
def import_data():
    if request.method == 'POST':
        # Validate idempotency token to prevent double-submission
        token = request.form.get('import_token', '')
        if token != session.pop('import_token', None):
            flash('Formulario ya fue procesado. Por favor intente nuevamente.', 'warning')
            return redirect(url_for('import_data'))

        if 'file' not in request.files:
            flash('No se seleccionó archivo.', 'danger')
            return redirect(url_for('import_data'))

        file = request.files['file']
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Solo se aceptan archivos Excel (.xlsx, .xls).', 'danger')
            return redirect(url_for('import_data'))

        try:
            from openpyxl import load_workbook
            from sqlalchemy.exc import IntegrityError
            import unicodedata
            import tempfile
            fd, filepath = tempfile.mkstemp(suffix='.xlsx')
            os.close(fd)
            file.save(filepath)
            wb = load_workbook(filepath, read_only=True)

            # Acquire advisory lock to prevent concurrent imports (PostgreSQL only)
            if db.engine.dialect.name == 'postgresql':
                db.session.execute(db.text("SELECT pg_advisory_xact_lock(42)"))

            # Find the data sheet
            ws = None
            for name in wb.sheetnames:
                if 'registro' in name.lower() or 'documento' in name.lower():
                    ws = wb[name]
                    break
            if ws is None:
                ws = wb[wb.sheetnames[0]]

            imported = 0
            skipped = 0
            updated = 0
            archivados = 0
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                flash('El archivo está vacío.', 'warning')
                return redirect(url_for('import_data'))

            headers = [str(h).strip().upper() if h else '' for h in rows[0]]

            # Reject files that don't look like an employee sheet (e.g. an
            # equipment template uploaded by mistake creates ghost PLI rows).
            def _norm(s):
                return unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode()
            normalized_headers = [_norm(h) for h in headers]
            joined_headers = ' '.join(normalized_headers)
            first_header = normalized_headers[0] if normalized_headers else ''
            equipment_markers = ('TITULAR', 'UNIDAD', 'VIN', 'MARBETE', 'TABLILLA')
            looks_like_employee = ('NOMBRE' in first_header) or ('NAME' in first_header)
            looks_like_equipment = any(m in joined_headers for m in equipment_markers)
            if not looks_like_employee or looks_like_equipment:
                wb.close()
                os.remove(filepath)
                flash('El archivo no parece ser de empleados. Verifique que está usando la plantilla correcta.', 'danger')
                return redirect(url_for('import_data'))

            for row in rows[1:]:
                if not row[0]:
                    continue

                name = str(row[0]).strip()

                # Check if employee already exists
                existing = Employee.query.filter_by(name=name).first()
                if existing and existing.is_archived:
                    # Archived employees are frozen — a stale spreadsheet must
                    # not keep maintaining their permits.
                    archivados += 1
                    continue
                if existing:
                    # Update permits for existing employee
                    permit_map = {
                        7: 'NTSP',
                        8: 'TWIC',
                        9: 'CERT_MEDICO',
                        10: 'ANTECEDENTES',
                        11: 'RECORD_CHOFERIL',
                        13: 'HM126',
                        14: 'HM232',
                        15: 'PRIMEROS_AUXILIOS',
                    }
                    for col_idx, ptype in permit_map.items():
                        if col_idx >= len(row):
                            continue
                        val = row[col_idx]
                        applicability = 'YES'
                        exp_date = None

                        if val is None:
                            applicability = 'YES'
                            exp_date = None
                        elif isinstance(val, str):
                            val_upper = val.strip().upper()
                            if val_upper in ('N/A', 'NA', ''):
                                applicability = 'N/A'
                            elif val_upper in ('NO',):
                                applicability = 'N/A'
                        elif hasattr(val, 'date'):
                            d = val.date() if hasattr(val, 'date') else val
                            if hasattr(d, 'year') and d.year < 2000:
                                try:
                                    d = d.replace(year=d.year + 100)
                                except ValueError:
                                    pass
                            exp_date = d

                        permit = EmployeePermit.query.filter_by(
                            employee_id=existing.id, permit_type=ptype
                        ).first()
                        if permit:
                            permit.applicability = applicability
                            permit.expiration_date = exp_date
                        else:
                            db.session.add(EmployeePermit(
                                employee_id=existing.id,
                                permit_type=ptype,
                                applicability=applicability,
                                expiration_date=exp_date,
                            ))
                    updated += 1
                    continue

                # Map company
                company_raw = str(row[1]).strip().upper() if row[1] else ''
                company = 'LB' if 'LB' in company_raw else 'PLI'

                area = str(row[2]).strip() if row[2] else ''
                status = str(row[3]).strip() if row[3] else 'activo'

                # License number
                lic_num = str(int(row[4])) if row[4] and isinstance(row[4], (int, float)) else (str(row[4]).strip() if row[4] else '')

                # License expiration - fix 2-digit year issue
                lic_exp = None
                if row[5] and hasattr(row[5], 'date'):
                    d = row[5].date() if hasattr(row[5], 'date') else row[5]
                    # Fix dates with year < 2000 (Excel 2-digit year bug)
                    if hasattr(d, 'year') and d.year < 2000:
                        try:
                            d = d.replace(year=d.year + 100)
                        except ValueError:
                            pass
                    lic_exp = d

                # Shirt size (column 12, index 12)
                shirt = str(row[12]).strip() if len(row) > 12 and row[12] else ''

                # Endoso HAZMAT (column 6)
                hazmat_raw = str(row[6]).strip().upper() if row[6] else 'N/A'
                if hazmat_raw in ('SI', 'SÍ', 'YES'):
                    hazmat = 'SI'
                elif hazmat_raw in ('NO',):
                    hazmat = 'NO'
                else:
                    hazmat = 'N/A'

                emp = Employee(
                    name=name,
                    company=company,
                    area=area,
                    status=status,
                    license_number=lic_num,
                    license_expiration=lic_exp,
                    shirt_size=shirt,
                    endoso_hazmat=hazmat,
                )
                db.session.add(emp)
                db.session.flush()

                # Create permits from columns
                permit_map = {
                    7: 'NTSP',
                    8: 'TWIC',
                    9: 'CERT_MEDICO',
                    10: 'ANTECEDENTES',
                    11: 'RECORD_CHOFERIL',
                    13: 'HM126',
                    14: 'HM232',
                    15: 'PRIMEROS_AUXILIOS',
                }

                for col_idx, ptype in permit_map.items():
                    if col_idx >= len(row):
                        continue
                    val = row[col_idx]
                    applicability = 'YES'
                    exp_date = None

                    if val is None:
                        applicability = 'YES'  # Applicable but missing
                        exp_date = None
                    elif isinstance(val, str):
                        val_upper = val.strip().upper()
                        if val_upper in ('N/A', 'NA', ''):
                            applicability = 'N/A'
                        elif val_upper in ('NO',):
                            applicability = 'N/A'
                    elif hasattr(val, 'date'):
                        d = val.date() if hasattr(val, 'date') else val
                        if hasattr(d, 'year') and d.year < 2000:
                            try:
                                d = d.replace(year=d.year + 100)
                            except ValueError:
                                pass
                        exp_date = d

                    permit = EmployeePermit(
                        employee_id=emp.id,
                        permit_type=ptype,
                        applicability=applicability,
                        expiration_date=exp_date,
                    )
                    db.session.add(permit)

                imported += 1

            try:
                db.session.commit()
            except IntegrityError:
                db.session.rollback()
                flash('Error: datos duplicados detectados. Posible importación simultánea.', 'danger')
                os.remove(filepath)
                return redirect(url_for('import_data'))
            os.remove(filepath)
            archived_clause = f', {archivados} omitidos por archivado' if archivados else ''
            flash(f'Importación completa: {imported} nuevos, {updated} actualizados, {skipped} omitidos{archived_clause}.', 'success')

        except ProtectedFileError as e:
            db.session.rollback()
            flash(f'Importación cancelada para proteger documentos adjuntos: {e}', 'danger')
        except Exception as e:
            db.session.rollback()
            flash(f'Error en la importación: {str(e)}', 'danger')

        return redirect(url_for('import_data'))

    session['import_token'] = secrets.token_hex(16)
    session['issue_import_token'] = secrets.token_hex(16)
    return render_template('import.html',
                           import_token=session['import_token'],
                           issue_import_token=session['issue_import_token'],
                           categories=ISSUE_CATEGORIES,
                           severities=ISSUE_SEVERITIES,
                           statuses=ISSUE_STATUSES)


@app.route('/import/equipment', methods=['POST'])
@admin_required
@protect_uploaded_files()   # documents are manual-only; see models.py
def import_equipment():
    """Upsert the equipment master sheet.

    Columns are located by HEADER NAME (see EQUIPMENT_COLUMN_ALIASES), never by
    position: a spreadsheet that gains or reorders a column used to shift every
    field silently — models landing in vin_serial, years in plate_number, permit
    dates blanked, and rows whose Compañía cell moved out from under the reader
    defaulting into LB. Rows whose company cannot be resolved are reported, not
    filed under a default company.
    """

    def to_date(cell):
        """Coerce an Excel cell to a date or None."""
        if cell is None or cell == '':
            return None
        if isinstance(cell, datetime):
            return cell.date()
        if isinstance(cell, date):
            return cell
        if isinstance(cell, str):
            s = cell.strip()
            if not s:
                return None
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
                try:
                    return datetime.strptime(s, fmt).date()
                except ValueError:
                    continue
        return None

    def to_float(cell):
        if cell is None or cell == '':
            return None
        try:
            return float(cell)
        except (ValueError, TypeError):
            return None

    # Validate idempotency token
    token = request.form.get('import_token', '')
    if token != session.pop('import_token', None):
        flash('Formulario ya fue procesado. Por favor intente nuevamente.', 'warning')
        return redirect(url_for('import_data'))

    if 'file' not in request.files:
        flash('No se seleccionó archivo.', 'danger')
        return redirect(url_for('import_data'))

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        flash('Solo se aceptan archivos Excel (.xlsx, .xls).', 'danger')
        return redirect(url_for('import_data'))

    try:
        from openpyxl import load_workbook
        from sqlalchemy.exc import IntegrityError

        import tempfile
        fd, filepath = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        file.save(filepath)
        wb = load_workbook(filepath, read_only=True)

        # Advisory lock for PostgreSQL
        if db.engine.dialect.name == 'postgresql':
            db.session.execute(db.text("SELECT pg_advisory_xact_lock(43)"))

        # Find data sheet
        ws = None
        for name in wb.sheetnames:
            if 'equipo' in name.lower() or 'vehicle' in name.lower():
                ws = wb[name]
                break
        if ws is None:
            ws = wb[wb.sheetnames[0]]

        imported = 0
        skipped = 0
        archivados = 0
        sin_empresa = 0          # rows whose Compañía cell could not be read
        reasignados = 0          # rows moved to the company the sheet declares
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            flash('El archivo está vacío.', 'warning')
            return redirect(url_for('import_data'))

        # Reject files that don't look like an equipment sheet (mirrors the
        # guard in import_data() so neither endpoint accepts the wrong template).
        normalized_headers = [_normalize_header(h) for h in rows[0]]
        joined_headers = ' '.join(normalized_headers)
        employee_markers = ('TWIC', 'HAZMAT', 'CERT MEDICO', 'ANTECEDENTES')
        if any(m in joined_headers for m in employee_markers):
            wb.close()
            os.remove(filepath)
            flash('El archivo no parece ser de equipos. Verifique que está usando la plantilla correcta.', 'danger')
            return redirect(url_for('import_data'))

        columns = _map_equipment_columns(normalized_headers)
        missing = _missing_equipment_columns(columns)
        if missing:
            wb.close()
            os.remove(filepath)
            found = ', '.join(h for h in normalized_headers if h) or '(ninguno)'
            flash(
                'No se pudieron identificar estas columnas en el archivo: '
                f'{", ".join(missing)}. Encabezados encontrados: {found}. '
                'Revise que la primera fila tenga los títulos de columna.',
                'danger',
            )
            return redirect(url_for('import_data'))

        def parse_permit_cell(val):
            if val is None:
                return ('YES', None)
            if isinstance(val, str):
                val_upper = val.strip().upper()
                if val_upper in ('N/A', 'NA', '', 'NO'):
                    return ('N/A', None)
                parsed = to_date(val)
                return ('YES', parsed)
            if hasattr(val, 'date'):
                d = val.date() if callable(getattr(val, 'date', None)) else val
                if hasattr(d, 'year') and d.year < 2000:
                    try:
                        d = d.replace(year=d.year + 100)
                    except ValueError:
                        pass
                return ('YES', d)
            return ('YES', to_date(val))

        def text_at(row, field):
            """Trimmed string from the column mapped to `field` ('' when absent)."""
            idx = columns.get(field)
            if idx is None or idx >= len(row) or row[idx] is None:
                return ''
            return str(row[idx]).strip()

        def digits_only(value):
            """Strip Excel's trailing .0 from numeric-looking identifiers."""
            if value and value.replace('.', '').replace(',', '').isdigit():
                return str(int(float(value.replace(',', ''))))
            return value

        for row in rows[1:]:
            if not row or not any(row):
                continue

            titulo = text_at(row, 'titular')
            unit_number = digits_only(text_at(row, 'unit'))

            # The sheet is the master record: an unreadable company is reported,
            # never silently filed under a default one (that is what put PLI and
            # Personal vehicles in the LB list).
            company = _resolve_company(text_at(row, 'company'), titulo)

            model = text_at(row, 'model')
            make = text_at(row, 'make')

            year = None
            year_raw = row[columns['year']] if columns.get('year') is not None and columns['year'] < len(row) else None
            if year_raw is not None:
                try:
                    year = int(float(str(year_raw).strip()))
                except (ValueError, TypeError):
                    pass

            vin_serial = digits_only(text_at(row, 'vin'))
            plate_number = digits_only(text_at(row, 'plate'))

            # Skip rows with no identifying data (empty/junk rows from Excel)
            if not titulo and not unit_number and not vin_serial and not plate_number:
                continue

            if company is None:
                sin_empresa += 1
                continue

            insurance_company = text_at(row, 'insurance')
            equipment_class = _resolve_import_class(text_at(row, 'clase'), model)

            # Only permits the sheet actually carries are touched — a missing
            # column must never blank a stored expiration date.
            permit_info = {}
            for ptype in ('MARBETE', 'INSPECCION', 'NTSP'):
                idx = columns.get(ptype)
                if idx is None:
                    continue
                permit_info[ptype] = parse_permit_cell(row[idx] if idx < len(row) else None)

            cost = to_float(row[columns['cost']]) if columns.get('cost') is not None and columns['cost'] < len(row) else None

            # Look up existing equipment via the shared case-insensitive cascade
            # (VIN → plate → unit#+company) and upsert in place, so re-imports never
            # create duplicates — even when VIN/plate are blank but the unit# is known.
            existing = find_duplicate_equipment(vin_serial, plate_number, unit_number, company,
                                                titular=titulo, model=model, year=year)

            if existing and existing.is_archived:
                # Archived equipment is frozen — skip the row instead of
                # refreshing its permits (or minting a duplicate that would
                # trip the partial unique indexes).
                archivados += 1
                continue

            if existing:
                if titulo:            existing.titular = titulo
                if insurance_company: existing.insurance_company = insurance_company
                if cost is not None:  existing.cost = cost
                if year is not None:  existing.year = year
                if model:             existing.model = model
                if make:              existing.make = make
                if existing.company != company:
                    # The sheet reassigned the vehicle; without this an earlier
                    # bad import could never be corrected by re-importing.
                    existing.company = company
                    reasignados += 1
                existing.equipment_class = equipment_class

                existing_permits = {p.permit_type: p for p in existing.permits}
                for ptype, (applicability, pdate) in permit_info.items():
                    if ptype in existing_permits:
                        permit = existing_permits[ptype]
                        permit.applicability = applicability
                        permit.expiration_date = pdate if applicability == 'YES' else None
                    else:
                        db.session.add(EquipmentPermit(
                            equipment_id=existing.id,
                            permit_type=ptype,
                            applicability=applicability,
                            expiration_date=pdate if applicability == 'YES' else None,
                        ))
                # The sheet can reassign company or class, which flips whether a
                # shared policy covers the vehicle — realign its SEGURO permit.
                sync_vehicle_insurance_permit(existing)
                skipped += 1
                continue

            equip = Equipment(
                company=company,
                titular=titulo,
                unit_number=unit_number,
                make=make,
                model=model,
                equipment_class=equipment_class,
                year=year,
                vin_serial=vin_serial,
                plate_number=plate_number,
                insurance_company=insurance_company,
                cost=cost,
                status='activo',
            )
            db.session.add(equip)
            db.session.flush()

            for ptype, (applicability, pdate) in permit_info.items():
                db.session.add(EquipmentPermit(
                    equipment_id=equip.id,
                    permit_type=ptype,
                    applicability=applicability,
                    expiration_date=pdate if applicability == 'YES' else None,
                ))
            # The sheet carries no per-vehicle insurance column, so this creates the
            # SEGURO slot at the right applicability for the vehicle's policy.
            sync_vehicle_insurance_permit(equip)

            imported += 1

        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash('Error: datos duplicados detectados. Posible importación simultánea.', 'danger')
            os.remove(filepath)
            return redirect(url_for('import_data'))

        os.remove(filepath)
        extra = ''
        if archivados:
            extra += f', {archivados} omitidos por archivado'
        if reasignados:
            extra += f', {reasignados} reasignados de empresa'
        flash(f'Importación completa: {imported} equipos importados, {skipped} actualizados{extra}.', 'success')
        if sin_empresa:
            flash(
                f'{sin_empresa} fila(s) se omitieron porque la columna Compañía está vacía o no '
                'se reconoce (valores válidos: LB, PLI, Personal). Complete esa columna y vuelva a importar.',
                'warning',
            )

    except ProtectedFileError as e:
        db.session.rollback()
        flash(f'Importación cancelada para proteger documentos adjuntos: {e}', 'danger')
    except Exception as e:
        db.session.rollback()
        flash(f'Error en la importación: {str(e)}', 'danger')

    return redirect(url_for('import_data'))


# ── PDF REPORT ─────────────────────────────────────────────────────────

@app.route('/reports')
@login_required
def report_menu():
    """Picker page: choose report purpose + company, then download the scoped PDF."""
    return render_template('report_options.html')


def _gather_report_data(report_type, company_filter):
    """Shared querysets for the PDF and Excel reports (archived excluded,
    open issues only)."""
    employees = []
    equipment_list = []
    issues = []
    company_permits = []

    if report_type in ('all', 'employees'):
        q = (Employee.query
             .filter(Employee.archived_at.is_(None))
             .order_by(Employee.company, Employee.name))
        if company_filter:
            q = q.filter(Employee.company == company_filter)
        employees = q.all()

    if report_type in ('all', 'equipment'):
        q = (Equipment.query
             .filter(Equipment.archived_at.is_(None))
             .order_by(Equipment.company, Equipment.make))
        if company_filter:
            q = q.filter(Equipment.company == company_filter)
        equipment_list = q.all()

    if report_type in ('all', 'issues'):
        # Open issues only; grouped by company via the linked Equipment.
        # Issues on archived (sold/scrapped) vehicles are dead work items —
        # excluded here, still visible in-app on the equipment detail page.
        iq = (Issue.query
              .join(Equipment, Issue.equipment_id == Equipment.id)
              .filter(Issue.current_status.notin_(RESOLVED_STATUSES),
                      Equipment.archived_at.is_(None)))
        if company_filter:
            iq = iq.filter(Equipment.company == company_filter)
        # Sort in Python: company (matches the template's LB/PLI grouping), then
        # by vehicle (unit # / display name), then newest first within a vehicle.
        issues = sorted(
            iq.all(),
            key=lambda i: (
                i.equipment.company if i.equipment else '',
                i.equipment.display_name.lower() if i.equipment else '',
                -(i.reported_at.toordinal() if i.reported_at else 0),
            ),
        )

    if report_type in ('all', 'company'):
        cq = CompanyPermit.query.order_by(CompanyPermit.company, CompanyPermit.permit_type)
        if company_filter:
            cq = cq.filter(CompanyPermit.company == company_filter)
        company_permits = cq.all()

    return employees, equipment_list, issues, company_permits


def _insurance_permits_by_equipment(equipment_list):
    """equipment id → the permit record covering its insurance.

    That is the shared company policy for vehicles on one, else the vehicle's own
    SEGURO permit. Shared by generate_pdf and export_excel so the Seguro column
    cannot mean different things in the two formats.
    """
    shared_policies = {
        (cp.company, cp.permit_type): cp
        for cp in CompanyPermit.query.filter(
            CompanyPermit.permit_type.in_(list(INSURANCE_TYPE_BY_CLASS.values()))).all()
    }
    resolved = {}
    for eq in equipment_list:
        ptype = eq.insurance_permit_type
        resolved[eq.id] = (shared_policies.get((eq.company, ptype)) if ptype
                           else eq.permits.filter_by(permit_type='SEGURO').first())
    return resolved


@app.route('/report/pdf')
@login_required
def generate_pdf():
    report_type = request.args.get('type', 'all')  # all, employees, equipment, issues, company
    company_filter = request.args.get('company', '')

    employees, equipment_list, issues, company_permits = _gather_report_data(
        report_type, company_filter)

    html = render_template('report_pdf.html',
        employees=employees,
        equipment_list=equipment_list,
        insurance_by_eq=_insurance_permits_by_equipment(equipment_list),
        issues=issues,
        company_permits=company_permits,
        report_type=report_type,
        today=date.today(),
        alert_date=date.today() + timedelta(days=30),
    )

    try:
        from weasyprint import HTML
        pdf = HTML(string=html, base_url=request.url_root).write_pdf()
        from io import BytesIO
        buffer = BytesIO(pdf)
        buffer.seek(0)
        fname = f"reporte_{report_type}_{company_filter or 'todas'}_{date.today().strftime('%Y%m%d')}.pdf"
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=fname)
    except Exception as e:
        # WeasyPrint needs native libs (Pango/Cairo/gobject) that may be absent
        # in local/dev environments, where the import raises OSError (not just
        # ImportError). Degrade gracefully to a printable HTML report instead of
        # 500-ing the route.
        app.logger.warning('PDF generation unavailable, serving HTML fallback: %s', e)
        return html


# Excel status colors mirror the light-theme status tokens in style.css.
_XLSX_STATUS_COLORS = {
    'valid': ('DCFCE7', '166534'),
    'expiring_soon': ('FEF3C7', '92400E'),
    'expired': ('FEE2E2', '991B1B'),
    'missing': ('F1F5F9', '64748B'),
}
_XLSX_SEVERITY_COLORS = {
    'baja': ('DBEAFE', '1D4ED8'),
    'media': ('FEF3C7', '92400E'),
    'alta': ('FED7AA', 'C2410C'),
    'critica': ('FECACA', 'DC2626'),
}


@app.route('/export/excel')
@login_required
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    report_type = request.args.get('type', 'all')
    company_filter = request.args.get('company', '')

    employees, equipment_list, issues, company_permits = _gather_report_data(
        report_type, company_filter)

    today = date.today()
    alert_date = today + timedelta(days=30)

    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(fill_type='solid', start_color='E2E8F0')

    def style_status(cell, status):
        colors = _XLSX_STATUS_COLORS.get(status)
        if colors:
            cell.fill = PatternFill(fill_type='solid', start_color=colors[0])
            cell.font = Font(color=colors[1])

    def date_status(d):
        if d is None:
            return 'missing'
        if d < today:
            return 'expired'
        if d <= alert_date:
            return 'expiring_soon'
        return 'valid'

    def write_permit_cell(ws, row, col, permit):
        cell = ws.cell(row=row, column=col)
        if permit is None:
            cell.value = '—'
            style_status(cell, 'missing')
        elif permit.applicability == 'N/A':
            cell.value = 'N/A'
        elif permit.expiration_date is None:
            cell.value = '—'
            style_status(cell, 'missing')
        else:
            cell.value = permit.expiration_date
            cell.number_format = 'MM/DD/YYYY'
            style_status(cell, permit.status)
        return cell

    def write_date_cell(ws, row, col, d):
        cell = ws.cell(row=row, column=col)
        if d is None:
            cell.value = '—'
            style_status(cell, 'missing')
        else:
            cell.value = d
            cell.number_format = 'MM/DD/YYYY'
            style_status(cell, date_status(d))
        return cell

    def new_sheet(wb, title, headers, widths):
        ws = wb.create_sheet(title)
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        ws.freeze_panes = 'A2'
        return ws

    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet; we create named ones

    if employees:
        emp_headers = ['Empresa', 'Nombre', 'Área', 'Puesto', 'Licencia'] + \
                      [label for _code, label in EMPLOYEE_PERMIT_TYPES if _code != 'OTHER']
        ws = new_sheet(wb, 'Empleados', emp_headers,
                       [10, 28, 16, 16, 12] + [14] * (len(emp_headers) - 5))
        for r, emp in enumerate(employees, start=2):
            ws.cell(row=r, column=1, value=emp.company)
            ws.cell(row=r, column=2, value=emp.name)
            ws.cell(row=r, column=3, value=emp.area or '')
            ws.cell(row=r, column=4, value=emp.puesto or '')
            write_date_cell(ws, r, 5, emp.license_expiration)
            permits_by_type = {p.permit_type: p for p in emp.permits}
            col = 6
            for code, _label in EMPLOYEE_PERMIT_TYPES:
                if code == 'OTHER':
                    continue
                write_permit_cell(ws, r, col, permits_by_type.get(code))
                col += 1

    if equipment_list:
        insurance_by_eq = _insurance_permits_by_equipment(equipment_list)
        eq_headers = ['Empresa', 'Unidad', 'Titular', 'Clase', 'Marca', 'Modelo', 'Año',
                      'VIN', 'Tablilla', 'Marbete', 'Seguro', 'Inspección', 'NTSP']
        ws = new_sheet(wb, 'Equipos', eq_headers,
                       [10, 12, 22, 12, 14, 18, 8, 20, 12, 14, 14, 14, 14])
        for r, eq in enumerate(equipment_list, start=2):
            ws.cell(row=r, column=1, value=eq.company)
            ws.cell(row=r, column=2, value=eq.display_name)
            ws.cell(row=r, column=3, value=eq.titular or '')
            ws.cell(row=r, column=4, value=eq.equipment_class_label)
            ws.cell(row=r, column=5, value=eq.make or '')
            ws.cell(row=r, column=6, value=eq.model or '')
            ws.cell(row=r, column=7, value=eq.year)
            ws.cell(row=r, column=8, value=eq.vin_serial or '')
            ws.cell(row=r, column=9, value=eq.plate_number or '')
            permits_by_type = {p.permit_type: p for p in eq.permits}
            write_permit_cell(ws, r, 10, permits_by_type.get('MARBETE'))
            write_permit_cell(ws, r, 11, insurance_by_eq.get(eq.id))
            write_permit_cell(ws, r, 12, permits_by_type.get('INSPECCION'))
            write_permit_cell(ws, r, 13, permits_by_type.get('NTSP'))

    if issues:
        ws = new_sheet(wb, 'Averías',
                       ['Empresa', 'Unidad', 'Categoría', 'Severidad', 'Estado',
                        'Descripción', 'Reportado por', 'Fecha Reporte'],
                       [10, 12, 16, 12, 14, 50, 22, 14])
        for r, issue in enumerate(issues, start=2):
            eq = issue.equipment
            ws.cell(row=r, column=1, value=eq.company if eq else '')
            ws.cell(row=r, column=2, value=eq.display_name if eq else '')
            ws.cell(row=r, column=3, value=issue.category_label)
            sev = ws.cell(row=r, column=4, value=issue.severity_label)
            colors = _XLSX_SEVERITY_COLORS.get(issue.severity)
            if colors:
                sev.fill = PatternFill(fill_type='solid', start_color=colors[0])
                sev.font = Font(color=colors[1])
            ws.cell(row=r, column=5, value=issue.status_label)
            ws.cell(row=r, column=6, value=issue.description or '')
            ws.cell(row=r, column=7, value=issue.reporter.name if issue.reporter else '')
            if issue.reported_at:
                d = ws.cell(row=r, column=8, value=issue.reported_at.date())
                d.number_format = 'MM/DD/YYYY'

    if company_permits:
        ws = new_sheet(wb, 'Empresa',
                       ['Empresa', 'Permiso', 'Vence', 'Número', 'Autoridad'],
                       [10, 28, 14, 20, 24])
        for r, cp in enumerate(company_permits, start=2):
            ws.cell(row=r, column=1, value=cp.company)
            ws.cell(row=r, column=2, value=cp.display_name)
            write_permit_cell(ws, r, 3, cp)
            ws.cell(row=r, column=4, value=cp.permit_number or '')
            ws.cell(row=r, column=5, value=cp.issuing_authority or '')

    if not wb.sheetnames:
        wb.create_sheet('Reporte')  # avoid an invalid zero-sheet workbook

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    fname = f"reporte_{report_type}_{company_filter or 'todas'}_{today.strftime('%Y%m%d')}.xlsx"
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=fname)


# ── USER MANAGEMENT ────────────────────────────────────────────────────

@app.route('/admin/users')
@admin_required
def admin_users():
    users = User.query.order_by(User.role, User.username).all()
    return render_template('admin_users.html', users=users)


@app.route('/admin/users/new', methods=['POST'])
@admin_required
def admin_user_new():
    username = request.form['username'].strip()
    email = request.form['email'].strip()
    password = request.form['password']
    role = request.form.get('role', 'viewer')

    if User.query.filter_by(username=username).first():
        flash(f'El usuario {username} ya existe.', 'danger')
        return redirect(url_for('admin_users'))

    user = User(username=username, email=email, role=role)
    user.set_password(password)
    db.session.add(user)
    db.session.flush()
    for role_name in request.form.getlist('issue_roles'):
        if role_name == 'shop':
            db.session.add(UserIssueRole(user_id=user.id, role=role_name))
    db.session.commit()
    flash(f'Usuario {username} creado como {role}.', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/users/<int:id>/edit', methods=['POST'])
@admin_required
def admin_user_edit(id):
    user = User.query.get_or_404(id)
    user.email = request.form.get('email', user.email)
    user.role = request.form.get('role', user.role)
    password = request.form.get('password', '')
    if password:
        user.set_password(password)
    UserIssueRole.query.filter_by(user_id=user.id).delete()
    for role_name in request.form.getlist('issue_roles'):
        if role_name == 'shop':
            db.session.add(UserIssueRole(user_id=user.id, role=role_name))
    db.session.commit()
    flash(f'Usuario {user.username} actualizado.', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/users/<int:id>/delete', methods=['POST'])
@admin_required
def admin_user_delete(id):
    user = User.query.get_or_404(id)
    if user.id == current_user.id:
        flash('No puede eliminarse a sí mismo.', 'danger')
        return redirect(url_for('admin_users'))
    db.session.delete(user)
    db.session.commit()
    flash(f'Usuario eliminado.', 'warning')
    return redirect(url_for('admin_users'))


# ── HEALTH CHECK ──────────────────────────────────────────────────────

@app.route('/health')
def health():
    try:
        db.session.execute(db.text('SELECT 1'))
        return jsonify({'status': 'ok', 'database': 'ok'})
    except Exception as e:
        app.logger.error(f"Health check DB failure: {e}")
        return jsonify({'status': 'degraded', 'database': str(e)}), 503


# ── DB INIT CLI ────────────────────────────────────────────────────────

@app.cli.command('init-db')
def init_db():
    """Create tables and default admin user."""
    db.create_all()
    if not User.query.filter_by(username='admin').first():
        admin = User(username='admin', email='admin@lbcaribe.com', role='admin')
        admin.set_password('admin123')
        db.session.add(admin)
        db.session.commit()
        print('Database initialized. Admin user created (admin / admin123)')
    else:
        print('Database already initialized.')


@app.cli.command('dedup')
@click.option('--dry-run', is_flag=True, help='Preview changes without modifying the database.')
def dedup_data(dry_run):
    """Remove duplicate employees and permits, keeping the record with the most data."""
    # --- Deduplicate employees by (name, company) ---
    dupes = (
        db.session.query(Employee.name, Employee.company, db.func.count(Employee.id))
        .group_by(Employee.name, Employee.company)
        .having(db.func.count(Employee.id) > 1)
        .all()
    )

    emp_deleted = 0
    for name, company, count in dupes:
        employees = Employee.query.filter_by(name=name, company=company).order_by(Employee.id).all()
        keeper = employees[0]
        for dup in employees[1:]:
            # Merge non-null fields from duplicate into keeper
            for col in ['area', 'status', 'fecha_nacimiento', 'license_number',
                        'license_expiration', 'license_file', 'puesto', 'telefono',
                        'email', 'fecha_contratacion', 'contacto_emergencia', 'shirt_size',
                        'archived_at', 'archive_reason', 'archive_note']:
                if getattr(keeper, col) is None and getattr(dup, col) is not None:
                    setattr(keeper, col, getattr(dup, col))
            if keeper.endoso_hazmat == 'N/A' and dup.endoso_hazmat != 'N/A':
                keeper.endoso_hazmat = dup.endoso_hazmat

            # Merge permit data: if keeper's permit is missing data that dup's has, copy it
            for dup_permit in dup.permits.all():
                keeper_permit = keeper.permits.filter_by(permit_type=dup_permit.permit_type).first()
                if keeper_permit:
                    if keeper_permit.expiration_date is None and dup_permit.expiration_date is not None:
                        keeper_permit.expiration_date = dup_permit.expiration_date
                    if keeper_permit.applicability == 'N/A' and dup_permit.applicability == 'YES':
                        keeper_permit.applicability = dup_permit.applicability
                    for field in ['permit_number', 'issuing_authority', 'file_path', 'renewal_cost', 'notes']:
                        if getattr(keeper_permit, field) is None and getattr(dup_permit, field) is not None:
                            setattr(keeper_permit, field, getattr(dup_permit, field))

            print(f'  {"[DRY RUN] " if dry_run else ""}Delete duplicate employee: {dup.name} ({dup.company}) id={dup.id}, keeping id={keeper.id}')
            if not dry_run:
                db.session.delete(dup)
            emp_deleted += 1

    # --- Deduplicate permits within each employee by (employee_id, permit_type) ---
    permit_dupes = (
        db.session.query(EmployeePermit.employee_id, EmployeePermit.permit_type, db.func.count(EmployeePermit.id))
        .filter(EmployeePermit.permit_type != 'OTHER')
        .group_by(EmployeePermit.employee_id, EmployeePermit.permit_type)
        .having(db.func.count(EmployeePermit.id) > 1)
        .all()
    )

    permit_deleted = 0
    for emp_id, ptype, count in permit_dupes:
        permits = EmployeePermit.query.filter_by(employee_id=emp_id, permit_type=ptype).order_by(EmployeePermit.id).all()
        keeper = permits[0]
        for dup in permits[1:]:
            # Merge non-null fields
            if keeper.expiration_date is None and dup.expiration_date is not None:
                keeper.expiration_date = dup.expiration_date
            if keeper.applicability == 'N/A' and dup.applicability == 'YES':
                keeper.applicability = dup.applicability
            for field in ['permit_number', 'issuing_authority', 'file_path', 'renewal_cost', 'notes']:
                if getattr(keeper, field) is None and getattr(dup, field) is not None:
                    setattr(keeper, field, getattr(dup, field))
            print(f'  {"[DRY RUN] " if dry_run else ""}Delete duplicate permit: employee_id={emp_id} type={ptype} id={dup.id}, keeping id={keeper.id}')
            if not dry_run:
                db.session.delete(dup)
            permit_deleted += 1

    # --- Same for equipment permits ---
    eq_permit_dupes = (
        db.session.query(EquipmentPermit.equipment_id, EquipmentPermit.permit_type, db.func.count(EquipmentPermit.id))
        .filter(EquipmentPermit.permit_type != 'OTHER')
        .group_by(EquipmentPermit.equipment_id, EquipmentPermit.permit_type)
        .having(db.func.count(EquipmentPermit.id) > 1)
        .all()
    )

    eq_permit_deleted = 0
    for eq_id, ptype, count in eq_permit_dupes:
        permits = EquipmentPermit.query.filter_by(equipment_id=eq_id, permit_type=ptype).order_by(EquipmentPermit.id).all()
        keeper = permits[0]
        for dup in permits[1:]:
            if keeper.expiration_date is None and dup.expiration_date is not None:
                keeper.expiration_date = dup.expiration_date
            if keeper.applicability == 'N/A' and dup.applicability == 'YES':
                keeper.applicability = dup.applicability
            for field in ['permit_number', 'issuing_authority', 'file_path', 'renewal_cost', 'notes']:
                if getattr(keeper, field) is None and getattr(dup, field) is not None:
                    setattr(keeper, field, getattr(dup, field))
            if not dry_run:
                db.session.delete(dup)
            eq_permit_deleted += 1

    # --- Deduplicate equipment (vehicles) by shared VIN / plate / unit#+company ---
    # Two rows are the same vehicle if they share ANY non-blank key. A single
    # cascade key would miss true duplicates where one copy has a VIN and the other
    # only a plate, so union rows that share any signature (union-find) instead.
    # VIN/plate are global; unit# is per-company (each company numbers its own units).
    equipment_list = Equipment.query.order_by(Equipment.id).all()
    parent = {e.id: e.id for e in equipment_list}

    def _find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def _union(a, b):
        ra, rb = _find(a), _find(b)
        if ra != rb:
            parent[max(ra, rb)] = min(ra, rb)

    sig_to_id = {}
    for equip in equipment_list:
        vin = (equip.vin_serial or '').strip().lower()
        plate = (equip.plate_number or '').strip().lower()
        unit = (equip.unit_number or '').strip().lower()
        sigs = []
        if vin:
            sigs.append(('vin', vin))
        if plate:
            sigs.append(('plate', plate))
        if unit:
            sigs.append(('unit', equip.company, unit))
        # Keyless fallback: rows with no VIN/plate/unit (carretones, generadores,
        # tanques, etc.) get no signature above and would never merge. Group exact
        # copies by (company, titular, model, year) when titular or model is present.
        if not sigs:
            titular = (equip.titular or '').strip().lower()
            model = (equip.model or '').strip().lower()
            if titular or model:
                sigs.append(('keyless', equip.company, titular, model, equip.year))
        for sig in sigs:
            if sig in sig_to_id:
                _union(equip.id, sig_to_id[sig])
            else:
                sig_to_id[sig] = equip.id

    equip_groups = {}
    for equip in equipment_list:
        equip_groups.setdefault(_find(equip.id), []).append(equip)

    equip_deleted = 0
    for root, group in equip_groups.items():
        if len(group) < 2:
            continue
        group.sort(key=lambda e: e.id)
        keeper = group[0]
        for dup in group[1:]:
            # Merge non-empty scalar fields into keeper where keeper's is empty.
            for col in ['name', 'titular', 'unit_number', 'plate_number', 'make', 'model',
                        'year', 'vin_serial', 'insurance_company', 'insurance_policy_type',
                        'cost', 'notes',
                        'equipment_type', 'status',
                        'archived_at', 'archive_reason', 'archive_note']:
                keeper_val = getattr(keeper, col)
                dup_val = getattr(dup, col)
                keeper_empty = keeper_val is None or keeper_val == ''
                dup_present = dup_val is not None and dup_val != ''
                if keeper_empty and dup_present:
                    setattr(keeper, col, dup_val)

            # Merge child permits: fill keeper's gaps, or re-point dup's permit to
            # keeper when keeper lacks that type (avoids cascade-deleting real data).
            for dup_permit in dup.permits.all():
                keeper_permit = keeper.permits.filter_by(permit_type=dup_permit.permit_type).first()
                if keeper_permit:
                    if keeper_permit.expiration_date is None and dup_permit.expiration_date is not None:
                        keeper_permit.expiration_date = dup_permit.expiration_date
                    if keeper_permit.applicability == 'N/A' and dup_permit.applicability == 'YES':
                        keeper_permit.applicability = dup_permit.applicability
                    for field in ['permit_number', 'issuing_authority', 'file_path', 'renewal_cost', 'notes']:
                        if getattr(keeper_permit, field) is None and getattr(dup_permit, field) is not None:
                            setattr(keeper_permit, field, getattr(dup_permit, field))
                elif not dry_run:
                    dup_permit.equipment_id = keeper.id

            # Reassign linked issue reports (FK is SET NULL, so they'd be orphaned).
            if not dry_run:
                Issue.query.filter_by(equipment_id=dup.id).update({'equipment_id': keeper.id})

            print(f'  {"[DRY RUN] " if dry_run else ""}Delete duplicate equipment: {dup.display_name} ({dup.company}) id={dup.id}, keeping id={keeper.id}')
            if not dry_run:
                db.session.flush()
                db.session.delete(dup)
            equip_deleted += 1

    # --- Deduplicate issue reports created by repeated bulk imports ---
    # Identity mirrors the importer: (equipment_id, category, normalized
    # description, canonical report day). Reuse the importer's helpers so the
    # two layers can never drift. _canonical_report_day collapses dateless rows
    # (stored at a volatile utcnow timestamp) regardless of import day, so the
    # duplicates that previously escaped both layers now merge. Child
    # IssueStatusHistory rows cascade-delete.
    from issues.routes import _canonical_report_day, _normalize_desc

    issue_groups = {}
    for issue in Issue.query.order_by(Issue.id).all():
        day = _canonical_report_day(issue.reported_at)
        key = (issue.equipment_id, issue.category, _normalize_desc(issue.description), day)
        issue_groups.setdefault(key, []).append(issue)

    issue_deleted = 0
    for key, group in issue_groups.items():
        if len(group) < 2:
            continue
        group.sort(key=lambda i: i.id)
        keeper = group[0]
        for dup in group[1:]:
            # Re-point any child issues referencing the doomed dup to the keeper.
            if not dry_run:
                Issue.query.filter_by(parent_issue_id=dup.id).update({'parent_issue_id': keeper.id})
            print(f'  {"[DRY RUN] " if dry_run else ""}Delete duplicate issue: equipment_id={dup.equipment_id} '
                  f'cat={dup.category} id={dup.id}, keeping id={keeper.id}')
            if not dry_run:
                db.session.flush()
                db.session.delete(dup)
            issue_deleted += 1

    if not dry_run:
        db.session.commit()

        # Apply unique constraints (db.create_all() won't add them to existing tables)
        dialect = db.engine.dialect.name
        if dialect == 'postgresql':
            constraints = [
                "ALTER TABLE employees ADD CONSTRAINT uq_employee_name_company UNIQUE (name, company)",
                "ALTER TABLE employee_permits ADD CONSTRAINT uq_employee_permit_type UNIQUE (employee_id, permit_type)",
                "ALTER TABLE equipment_permits ADD CONSTRAINT uq_equipment_permit_type UNIQUE (equipment_id, permit_type)",
            ]
            for sql in constraints:
                try:
                    db.session.execute(db.text(sql))
                    db.session.commit()
                    print(f'  Applied: {sql}')
                except Exception:
                    db.session.rollback()
                    print(f'  Constraint already exists or skipped: {sql.split("ADD CONSTRAINT ")[1].split(" ")[0]}')

            # Permanent backstop against duplicate vehicles: partial unique indexes
            # on non-blank keys (blanks excluded so unidentified rows can coexist).
            # VIN/plate global, unit# per-company — mirrors find_duplicate_equipment.
            equip_indexes = [
                ("uq_equipment_vin",
                 "CREATE UNIQUE INDEX IF NOT EXISTS uq_equipment_vin ON equipment (lower(vin_serial)) "
                 "WHERE vin_serial IS NOT NULL AND vin_serial <> ''"),
                ("uq_equipment_plate",
                 "CREATE UNIQUE INDEX IF NOT EXISTS uq_equipment_plate ON equipment (lower(plate_number)) "
                 "WHERE plate_number IS NOT NULL AND plate_number <> ''"),
                ("uq_equipment_company_unit",
                 "CREATE UNIQUE INDEX IF NOT EXISTS uq_equipment_company_unit ON equipment (company, lower(unit_number)) "
                 "WHERE unit_number IS NOT NULL AND unit_number <> ''"),
            ]
            for name, sql in equip_indexes:
                try:
                    db.session.execute(db.text(sql))
                    db.session.commit()
                    print(f'  Applied: {name}')
                except Exception as exc:
                    db.session.rollback()
                    print(f'  WARNING: could not create {name} — residual duplicates likely remain: {exc}')

    prefix = '[DRY RUN] ' if dry_run else ''
    print(f'{prefix}Dedup complete: {emp_deleted} duplicate employees, {permit_deleted} duplicate employee permits, {equip_deleted} duplicate equipment, {eq_permit_deleted} duplicate equipment permits, {issue_deleted} duplicate issues removed.')


def _migrate_issue_statuses(dry_run=False):
    """Rewrite retired issue-status codes to the merged three-status vocabulary.

    The shop workflow collapsed from five stages to three: en_revision and
    en_reparacion were the same working state (now en_proceso), and resuelto and
    cerrado the same terminal one (now cerrado). This rewrites
    ``issues.current_status`` plus both ``issue_status_history`` status columns,
    then drops transitions the merge made meaningless (a real
    en_revision → en_reparacion step becomes en_proceso → en_proceso, which reads
    as noise on the detail timeline). The ``from_status IS NULL`` creation rows
    are never touched.

    Idempotent — a second run finds nothing. Returns {label: rows_affected}.
    """
    targets = (('issues', 'current_status'),
               ('issue_status_history', 'from_status'),
               ('issue_status_history', 'to_status'))
    counts = {}

    for old, new in LEGACY_ISSUE_STATUSES.items():
        for table, column in targets:
            n = db.session.execute(
                db.text(f'SELECT COUNT(*) FROM {table} WHERE {column} = :old'),
                {'old': old}).scalar() or 0
            if n:
                counts[f'{table}.{column} {old} -> {new}'] = n
                db.session.execute(
                    db.text(f'UPDATE {table} SET {column} = :new WHERE {column} = :old'),
                    {'old': old, 'new': new})

    # Counted after the updates land in the session (not committed yet) so a
    # --dry-run reports the redundant rows the merge *would* create, not just
    # any that already existed.
    redundant = ('FROM issue_status_history '
                 'WHERE from_status IS NOT NULL AND from_status = to_status')
    n = db.session.execute(db.text(f'SELECT COUNT(*) {redundant}')).scalar() or 0
    if n:
        counts['issue_status_history transiciones redundantes eliminadas'] = n
        db.session.execute(db.text(f'DELETE {redundant}'))

    if dry_run:
        db.session.rollback()
    else:
        db.session.commit()
    return counts


@app.cli.command('migrate-issue-statuses')
@click.option('--dry-run', is_flag=True, help='Preview changes without modifying the database.')
def migrate_issue_statuses(dry_run):
    """Merge the retired issue statuses into the current three-status flow."""
    prefix = '[DRY RUN] ' if dry_run else ''
    counts = _migrate_issue_statuses(dry_run=dry_run)
    if not counts:
        print(f'{prefix}No hay estados heredados que migrar.')
        return
    for label, n in counts.items():
        print(f'  {prefix}{label}: {n} fila(s)')
    print(f'{prefix}{sum(counts.values())} fila(s) afectada(s).')


@app.cli.command('reclassify-equipment')
@click.option('--dry-run', is_flag=True, help='Preview changes without modifying the database.')
def reclassify_equipment(dry_run):
    """Re-run the equipment classifier over all non-archived equipment."""
    prefix = '[DRY RUN] ' if dry_run else ''
    changed = 0
    for eq in Equipment.query.filter(Equipment.archived_at.is_(None)).all():
        new_class = classify_equipment(eq.model, eq.equipment_type)
        if eq.equipment_class != new_class:
            print(f'  {prefix}{eq.display_name} ({eq.company}) — "{eq.model or ""}" → {new_class}'
                  f' (antes: {eq.equipment_class or "sin clase"})')
            eq.equipment_class = new_class
            changed += 1
    if not dry_run and changed:
        db.session.commit()
    print(f'{prefix}{changed} equipo(s) reclasificado(s).')


@app.cli.command('migrate-class-insurance')
@click.option('--dry-run', is_flag=True, help='Preview changes without modifying the database.')
def migrate_class_insurance(dry_run):
    """Split the single shared SEGURO CompanyPermit into per-class policies.

    The existing SEGURO row becomes SEGURO_TRUCK (the currently-insured fleet);
    empty CHASSIS/TANK/GENERATOR slots are created. Idempotent."""
    prefix = '[DRY RUN] ' if dry_run else ''

    for company in ('LB', 'PLI'):
        old = CompanyPermit.query.filter_by(company=company, permit_type='SEGURO').first()
        truck = CompanyPermit.query.filter_by(company=company, permit_type='SEGURO_TRUCK').first()

        if old and not truck:
            print(f'  {prefix}{company}: SEGURO → SEGURO_TRUCK (id={old.id}, '
                  f'exp={old.expiration_date}, file={"yes" if old.file_path else "no"})')
            if not dry_run:
                old.permit_type = 'SEGURO_TRUCK'
        elif old and truck:
            # The Empresa view was visited post-deploy and auto-created a blank
            # SEGURO_TRUCK — merge the old row into it (blank fields only).
            print(f'  {prefix}{company}: merging SEGURO id={old.id} into SEGURO_TRUCK id={truck.id}')
            if not dry_run:
                for field in ('expiration_date', 'permit_number', 'issuing_authority',
                              'file_path', 'renewal_cost', 'notes'):
                    if not getattr(truck, field):
                        setattr(truck, field, getattr(old, field))
                truck.applicability = old.applicability or truck.applicability
                # file_path was transferred — do NOT delete_stored_file here
                db.session.delete(old)
        elif truck:
            print(f'  {prefix}{company}: ya migrado (SEGURO_TRUCK existe)')
        else:
            print(f'  {prefix}{company}: sin póliza SEGURO previa — se crearán registros vacíos')

        for code in ('SEGURO_TRUCK', 'SEGURO_CHASSIS', 'SEGURO_TANK', 'SEGURO_GENERATOR'):
            exists = CompanyPermit.query.filter_by(company=company, permit_type=code).first()
            if not exists and not (code == 'SEGURO_TRUCK' and old):
                print(f'  {prefix}{company}: creando registro {code}')
                if not dry_run:
                    db.session.add(CompanyPermit(company=company, permit_type=code,
                                                 applicability='YES'))

    if not dry_run:
        db.session.commit()

    print(f'{prefix}Migración de seguros por clase completa.')


@app.cli.command('migrate-shared-insurance')
@click.option('--dry-run', is_flag=True, help='Preview changes without modifying the database.')
def migrate_shared_insurance(dry_run):
    """Move per-vehicle SEGURO into shared CompanyPermit for LB/PLI, then hide the per-vehicle copies."""
    prefix = '[DRY RUN] ' if dry_run else ''

    for company in ('LB', 'PLI'):
        permits = (
            EquipmentPermit.query
            .join(Equipment, EquipmentPermit.equipment_id == Equipment.id)
            .filter(Equipment.company == company, EquipmentPermit.permit_type == 'SEGURO')
            .all()
        )

        # Pick the best source row: prefer one with an attached file, then the latest expiration.
        def _rank(p):
            return (1 if p.file_path else 0,
                    p.expiration_date or date.min)
        source = max(permits, key=_rank) if permits else None

        company_permit = CompanyPermit.query.filter_by(company=company, permit_type='SEGURO').first()
        if not company_permit:
            company_permit = CompanyPermit(company=company, permit_type='SEGURO', applicability='YES')
            if not dry_run:
                db.session.add(company_permit)

        if source:
            print(f'  {prefix}{company}: pre-fill shared insurance from equipment_permit id={source.id} '
                  f'(exp={source.expiration_date}, file={"yes" if source.file_path else "no"})')
            company_permit.applicability = 'YES'
            company_permit.expiration_date = source.expiration_date
            company_permit.permit_number = source.permit_number
            company_permit.issuing_authority = source.issuing_authority
            company_permit.file_path = source.file_path
        else:
            print(f'  {prefix}{company}: no per-vehicle insurance found — leaving shared record blank')

        hidden = 0
        for p in permits:
            if p.applicability != 'N/A':
                p.applicability = 'N/A'
                hidden += 1
        print(f'  {prefix}{company}: hid {hidden} per-vehicle insurance permit(s)')

    if not dry_run:
        db.session.commit()

    print(f'{prefix}Shared-insurance migration complete.')


@app.cli.command('send-notifications')
@click.option('--dry-run', is_flag=True, help='Preview without sending emails.')
@click.option('--employee-id', type=int, default=None, help='Send to a single employee (for testing).')
def send_notifications(dry_run, employee_id):
    """Send email notifications for expiring/expired permits."""
    print(f'{"[DRY RUN] " if dry_run else ""}Running notification cycle...')
    result = run_notification_cycle(dry_run=dry_run, employee_id=employee_id)
    for line in result['details']:
        print(line)
    prefix = '[DRY RUN] ' if dry_run else ''
    print(f'{prefix}Done: {result["sent"]} sent, {result["skipped"]} skipped, {result["failed"]} failed.')


@app.cli.command('generate-token')
@click.argument('employee_id', type=int)
def generate_token(employee_id):
    """Generate an issue-reporting access token for an employee."""
    emp = Employee.query.get(employee_id)
    if not emp:
        print(f'Error: Employee with ID {employee_id} not found.')
        return
    import uuid
    emp.access_token = str(uuid.uuid4())
    db.session.commit()
    print(f'Token generated for {emp.name} ({emp.company}):')
    print(f'  /issues/report/{emp.access_token}')


@app.cli.command('assign-issue-role')
@click.argument('username')
@click.argument('role', type=click.Choice(['shop']))
def assign_issue_role(username, role):
    """Assign an issue-module role (shop) to a user."""
    user = User.query.filter_by(username=username).first()
    if not user:
        print(f'Error: User "{username}" not found.')
        return
    existing = UserIssueRole.query.filter_by(user_id=user.id, role=role).first()
    if existing:
        print(f'User "{username}" already has the "{role}" role.')
        return
    db.session.add(UserIssueRole(user_id=user.id, role=role))
    db.session.commit()
    print(f'Assigned "{role}" role to user "{username}".')


# For Railway — auto-init on startup
with app.app_context():
    try:
        db.create_all()

        # Additive schema sync: db.create_all() does not ALTER existing tables,
        # so any column added to a model after the initial deploy is missing in
        # production. Compare each model's columns to the live table and ADD
        # whatever is missing. Idempotent and safe to run on every boot.
        if db.engine.dialect.name == 'postgresql':
            from sqlalchemy import inspect as sa_inspect
            from sqlalchemy.schema import CreateColumn
            inspector = sa_inspect(db.engine)
            for table in db.metadata.sorted_tables:
                if not inspector.has_table(table.name):
                    continue
                existing_cols = {c['name'] for c in inspector.get_columns(table.name)}
                for col in table.columns:
                    if col.name in existing_cols:
                        continue
                    try:
                        col_ddl = str(CreateColumn(col).compile(db.engine))
                        sql = f'ALTER TABLE {table.name} ADD COLUMN IF NOT EXISTS {col_ddl}'
                        db.session.execute(db.text(sql))
                        db.session.commit()
                        print(f"[INFO] Schema backfill: added {table.name}.{col.name}")
                    except Exception as e:
                        db.session.rollback()
                        print(f"[WARN] Schema backfill skipped {table.name}.{col.name}: {e}")

        # Rename CPR → PRIMEROS_AUXILIOS in existing records
        db.session.execute(db.text("UPDATE employee_permits SET permit_type = 'PRIMEROS_AUXILIOS' WHERE permit_type = 'CPR'"))
        db.session.commit()

        # Merge the retired five-status issue vocabulary into the current three.
        # Runs on boot (not just via `flask migrate-issue-statuses`) because
        # RESOLVED_STATUSES no longer lists 'resuelto' — an unmigrated DB would
        # show every historically resolved issue as open in the queue and in the
        # PDF/Excel reports until the command was run by hand. Idempotent.
        migrated = _migrate_issue_statuses()
        if migrated:
            print(f"[INFO] Issue status migration: {migrated}")

        # Backfill equipment_class for rows that predate the column (idempotent)
        unclassified = Equipment.query.filter(Equipment.equipment_class.is_(None)).all()
        for eq in unclassified:
            eq.equipment_class = classify_equipment(eq.model, eq.equipment_type)
        if unclassified:
            db.session.commit()
            print(f"[INFO] Classified {len(unclassified)} equipment rows")

        if not User.query.filter_by(role='admin').first():
            admin = User(username='admin', email='admin@lbcaribe.com', role='admin')
            admin.set_password('admin123')
            db.session.add(admin)
            db.session.commit()
    except Exception as e:
        print(f"[WARN] Database init failed: {e}")

    if app.config.get('ENABLE_SCHEDULER'):
        from apscheduler.schedulers.background import BackgroundScheduler
        scheduler = BackgroundScheduler(daemon=True)
        day = app.config['NOTIFICATION_DAY'][:3].lower()
        hour = app.config['NOTIFICATION_HOUR']

        def _scheduled_notifications():
            with app.app_context():
                run_notification_cycle()

        scheduler.add_job(_scheduled_notifications, 'cron', day_of_week=day, hour=hour, misfire_grace_time=3600)
        scheduler.start()
        import atexit
        atexit.register(scheduler.shutdown)
        app.logger.info(f"[SCHEDULER] Started — notifications every {app.config['NOTIFICATION_DAY']} at {hour}:00 UTC")


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
