import os
import uuid
import secrets
import tempfile
import unicodedata
from collections import defaultdict
from datetime import datetime, date
from flask import render_template, request, redirect, url_for, flash, abort, session
from flask_login import login_required, current_user
from sqlalchemy import func
from sqlalchemy.orm import joinedload

from issues import issues_bp
from issues.decorators import shop_required, admin_required, shop_manager_required
from models import (
    db, Employee, Equipment, EquipmentPermit, Issue, IssueStatusHistory,
    ISSUE_CATEGORIES, ISSUE_SEVERITIES, ISSUE_STATUSES,
    EXCLUDED_EQUIPMENT_MODELS, EQUIPMENT_PERMIT_TYPES,
    classify_equipment, protect_uploaded_files,
)

SEVERITY_RANK = {'critica': 4, 'alta': 3, 'media': 2, 'baja': 1}
RESOLVED_STATUSES = ['resuelto', 'cerrado']
QUEUE_SORTS = ('unit', 'recent', 'oldest', 'severity')


def _normalize_unit(raw):
    """Normalize a unit identifier for matching: handle numeric Excel cells,
    a leading '#', and case/whitespace."""
    if raw is None:
        return ''
    if isinstance(raw, float) and raw.is_integer():
        raw = int(raw)            # 204.0 -> 204, avoids "204.0"
    s = str(raw).strip()
    s = s.lstrip('#').strip()     # "#204" -> "204"
    return s.lower()


def _unreportable_reason(eq):
    """If equipment exists but won't appear in the driver report form, explain why.
    Returns a Spanish reason string, or None if the equipment is reportable."""
    if eq.archived_at is not None:
        return 'equipo archivado'
    if eq.company not in ('LB', 'PLI'):
        return f'empresa "{eq.company}"'
    if eq.equipment_class and eq.equipment_class != 'truck':
        return f'clase "{eq.equipment_class_label}"'
    if eq.status != 'activo':
        return f'estado "{eq.status}"'
    # NULL-class fallback (rows that predate the classifier backfill)
    if eq.model and eq.model.strip().lower() in EXCLUDED_EQUIPMENT_MODELS:
        return f'modelo "{eq.model}" excluido'
    return None


def reportable_equipment_query(company=None):
    query = Equipment.query.filter(
        Equipment.status == 'activo',
        Equipment.archived_at.is_(None),
        db.or_(Equipment.equipment_class == 'truck',
               Equipment.equipment_class.is_(None)),
        Equipment.company.in_(['LB', 'PLI']),
    )
    # NULL-class fallback: excluded-model loop only bites rows the backfill missed
    for model_name in EXCLUDED_EQUIPMENT_MODELS:
        query = query.filter(db.or_(
            Equipment.model == None,
            func.lower(Equipment.model) != model_name,
        ))
    if company:
        query = query.filter_by(company=company)
    return query.order_by(Equipment.unit_number)


def reportable_equipment_list(company=None):
    """Reportable trucks in natural unit-number order — the shared order for
    every shop-side truck list (dropdowns + queue cards)."""
    return sorted(reportable_equipment_query(company=company).all(),
                  key=lambda eq: eq.unit_sort_key)


def update_issue_status(issue, new_status, changed_by_user_id=None,
                        changed_by_employee_id=None, notes=None):
    old_status = issue.current_status

    history = IssueStatusHistory(
        issue_id=issue.id,
        from_status=old_status,
        to_status=new_status,
        changed_by_user_id=changed_by_user_id,
        changed_by_employee_id=changed_by_employee_id,
        notes=notes,
    )
    db.session.add(history)

    issue.current_status = new_status

    if new_status in ('resuelto', 'cerrado') and issue.resolved_at is None:
        issue.resolved_at = datetime.utcnow()
    elif new_status not in ('resuelto', 'cerrado'):
        issue.resolved_at = None


BACKLOG_NOTE = 'Importado del backlog'


def synthesize_status_history(issue, final_status, reported_at, resolved_at=None,
                              initial_note=None):
    """Backfill an issue's timeline for manual/bulk entry.

    Unlike ``update_issue_status`` (which stamps ``changed_at`` with the current
    time), this writes ``IssueStatusHistory`` rows with explicit timestamps so an
    imported issue reflects its real history. ``issue`` must already be flushed so
    it has an id. Sets ``current_status`` and, for resolved/closed issues,
    ``resolved_at``. Also aligns ``reported_at``/``created_at`` so the issue sorts
    correctly in the queue.
    """
    issue.reported_at = reported_at
    issue.created_at = reported_at

    db.session.add(IssueStatusHistory(
        issue_id=issue.id,
        from_status=None,
        to_status='reportado',
        changed_at=reported_at,
        notes=initial_note or BACKLOG_NOTE,
    ))

    if final_status != 'reportado':
        final_at = resolved_at or reported_at
        db.session.add(IssueStatusHistory(
            issue_id=issue.id,
            from_status='reportado',
            to_status=final_status,
            changed_at=final_at,
            notes=BACKLOG_NOTE,
        ))

    issue.current_status = final_status
    if final_status in ('resuelto', 'cerrado'):
        issue.resolved_at = resolved_at or reported_at
    else:
        issue.resolved_at = None


# ── DRIVER REPORT FLOW ───────────────────────────────────────────────

@issues_bp.route('/report/<token>', methods=['GET'])
def report(token):
    employee = Employee.query.filter_by(access_token=token).first_or_404()
    if employee.is_archived:
        return render_template('issues/link_inactive.html'), 410
    trucks = reportable_equipment_list(company=employee.company)
    return render_template('issues/report.html',
                           employee=employee,
                           trucks=trucks,
                           categories=ISSUE_CATEGORIES,
                           severities=ISSUE_SEVERITIES,
                           problem_rows=[{}])


@issues_bp.route('/report/<token>', methods=['POST'])
def report_submit(token):
    employee = Employee.query.filter_by(access_token=token).first_or_404()
    if employee.is_archived:
        return render_template('issues/link_inactive.html'), 410

    equipment_id = request.form.get('equipment_id', type=int)

    # Problem fields arrive as parallel arrays (name="category[]" etc.), one entry
    # per problem row. They stay index-aligned because getlist preserves DOM order.
    categories = request.form.getlist('category[]')
    severities = request.form.getlist('severity[]')
    descriptions = request.form.getlist('description[]')

    # Zip into rows, then drop fully-empty rows (a stray blank row must not block
    # the batch). A row counts as empty when both category and description are blank.
    raw_rows = []
    for i in range(max(len(categories), len(severities), len(descriptions))):
        raw_rows.append({
            'category': (categories[i] if i < len(categories) else '').strip(),
            'severity': (severities[i] if i < len(severities) else 'media').strip() or 'media',
            'description': (descriptions[i] if i < len(descriptions) else '').strip(),
        })
    problem_rows = [r for r in raw_rows
                    if r['category'] or r['description']]

    valid_categories = [c[0] for c in ISSUE_CATEGORIES]
    valid_severities = [s[0] for s in ISSUE_SEVERITIES]

    errors = []
    if not equipment_id:
        errors.append('Debe seleccionar un camión.')
    elif not reportable_equipment_query(company=employee.company).filter_by(id=equipment_id).first():
        errors.append('Camión no válido.')

    if not problem_rows:
        errors.append('Debe incluir al menos un problema.')

    for idx, row in enumerate(problem_rows, start=1):
        if row['category'] not in valid_categories:
            errors.append(f'Problema {idx}: categoría no válida.')
        if row['severity'] not in valid_severities:
            errors.append(f'Problema {idx}: severidad no válida.')
        if not row['description']:
            errors.append(f'Problema {idx}: debe incluir una descripción.')

    if errors:
        trucks = reportable_equipment_list(company=employee.company)
        for error in errors:
            flash(error, 'danger')
        return render_template('issues/report.html',
                               employee=employee,
                               trucks=trucks,
                               categories=ISSUE_CATEGORIES,
                               severities=ISSUE_SEVERITIES,
                               problem_rows=problem_rows or [{}]), 422

    # reported_at is intentionally NOT sourced from the request: it defaults to
    # the model's datetime.utcnow (see models.py). The driver form has no date
    # field and must never gain one — a future date can only enter via the
    # admin/import paths, which validate against _is_future_date.
    # Each problem row becomes its own independent Issue; commit once after the loop
    # (mirrors the batch idiom in import_issues).
    created_ids = []
    for row in problem_rows:
        issue = Issue(
            equipment_id=equipment_id,
            reported_by_employee_id=employee.id,
            category=row['category'],
            severity=row['severity'],
            description=row['description'],
        )
        db.session.add(issue)
        db.session.flush()

        update_issue_status(issue, 'reportado', changed_by_employee_id=employee.id,
                            notes='Reporte inicial del chofer')
        created_ids.append(issue.id)
    db.session.commit()

    return redirect(url_for('issues.report_success', token=token,
                            issue_ids=','.join(str(i) for i in created_ids)))


@issues_bp.route('/report/<token>/success', methods=['GET'])
def report_success(token):
    employee = Employee.query.filter_by(access_token=token).first_or_404()
    if employee.is_archived:
        return render_template('issues/link_inactive.html'), 410
    issue_ids = [int(part) for part in request.args.get('issue_ids', '').split(',')
                 if part.strip().isdigit()]
    return render_template('issues/report_success.html', token=token, issue_ids=issue_ids)


# ── SHOP QUEUE ────────────────────────────────────────────────────────

def _fold(s):
    """Lowercase + strip accents, for accent-insensitive matching."""
    return unicodedata.normalize('NFKD', s.lower()).encode('ascii', 'ignore').decode()


def _search_clause(search):
    """OR-filter for the queue's free-text search: description, reporter name,
    and truck identifiers via SQL ilike; category labels matched
    accent-insensitively in Python so 'electrico' finds 'Eléctrico'."""
    like = f'%{search}%'
    folded = _fold(search)
    cat_codes = [code for code, label in ISSUE_CATEGORIES
                 if folded in _fold(label) or folded in code]
    clauses = [
        Issue.description.ilike(like),
        Employee.name.ilike(like),
        Equipment.unit_number.ilike(like),
        Equipment.plate_number.ilike(like),
        Equipment.make.ilike(like),
        Equipment.model.ilike(like),
    ]
    if cat_codes:
        clauses.append(Issue.category.in_(cat_codes))
    return db.or_(*clauses)


@issues_bp.route('/', methods=['GET'])
@shop_required
def queue():
    # Issues on archived (sold/scrapped) equipment are hidden from the queue;
    # unlinked issues (equipment_id NULL) stay visible.
    not_archived = db.or_(Issue.equipment_id.is_(None),
                          Equipment.archived_at.is_(None))

    # Unknown filter values silently reset to defaults (dashboard pattern)
    status_filter = request.args.get('status', '')
    if status_filter not in {code for code, _ in ISSUE_STATUSES}:
        status_filter = ''
    severity_filter = request.args.get('severity', '')
    if severity_filter not in {code for code, _ in ISSUE_SEVERITIES}:
        severity_filter = ''
    category_filter = request.args.get('category', '')
    if category_filter not in {code for code, _ in ISSUE_CATEGORIES}:
        category_filter = ''
    equipment_filter = request.args.get('equipment_id', type=int)
    search = request.args.get('search', '').strip()
    sort_by = request.args.get('sort', 'unit')
    if sort_by not in QUEUE_SORTS:
        sort_by = 'unit'

    def scoped_query(exclude=None):
        """Every active filter applied except the axis named by `exclude`, so
        each control's count equals what clicking it would show."""
        q = (Issue.query
             .join(Equipment, Issue.equipment_id == Equipment.id, isouter=True)
             .filter(not_archived))
        if search:
            q = (q.join(Employee, Issue.reported_by_employee_id == Employee.id,
                        isouter=True)
                 .filter(_search_clause(search)))
        if exclude != 'status':
            if status_filter:
                q = q.filter(Issue.current_status == status_filter)
            else:
                q = q.filter(Issue.current_status.notin_(RESOLVED_STATUSES))
        if exclude != 'severity' and severity_filter:
            q = q.filter(Issue.severity == severity_filter)
        if equipment_filter:
            q = q.filter(Issue.equipment_id == equipment_filter)
        if category_filter:
            q = q.filter(Issue.category == category_filter)
        return q

    issue_order = (Issue.reported_at.asc() if sort_by == 'oldest'
                   else Issue.reported_at.desc())
    all_issues = (scoped_query()
                  .options(joinedload(Issue.reporter),
                           joinedload(Issue.equipment))
                  .order_by(issue_order)
                  .all())
    if sort_by == 'severity':
        # Stable sort keeps the newest-first SQL order as the tiebreak
        all_issues.sort(key=lambda i: -SEVERITY_RANK.get(i.severity, 0))

    issues_by_equipment = defaultdict(list)
    for issue in all_issues:
        if issue.equipment_id:
            issues_by_equipment[issue.equipment_id].append(issue)

    filters_active = bool(search or status_filter or severity_filter
                          or equipment_filter or category_filter)

    def build_truck_profiles(company):
        all_trucks = reportable_equipment_list(company=company)
        profiles = []
        for eq in all_trucks:
            issues = issues_by_equipment.get(eq.id, [])
            # Filtered views hide trucks with no matching issues; the bare
            # default view still lists the whole fleet.
            if filters_active and not issues:
                continue
            worst_rank = max((SEVERITY_RANK.get(i.severity, 0) for i in issues), default=0)
            worst_sev = next((s for s, r in SEVERITY_RANK.items() if r == worst_rank), None)
            profiles.append({
                'equipment': eq,
                'issues': issues,
                'issue_count': len(issues),
                'worst_severity': worst_sev,
                'worst_severity_rank': worst_rank,
                'auto_expand': len(issues) > 0,
            })
        # Default ('unit') sorts trucks in natural unit-number order,
        # matching the truck dropdown order. The other sorts reorder cards by
        # their issues (empty trucks last, by unit number among themselves).
        def name_key(p):
            return p['equipment'].unit_sort_key

        if sort_by == 'severity':
            profiles.sort(key=lambda p: (-p['worst_severity_rank'], name_key(p)))
        elif sort_by == 'recent':
            # datetime.max - <newest> = ascending-comparable "age", so the
            # alphabetical tiebreak isn't reversed by reverse=True
            profiles.sort(key=lambda p: (
                0 if p['issues'] else 1,
                datetime.max - max((i.reported_at or datetime.min for i in p['issues']),
                                   default=datetime.min),
                name_key(p)))
        elif sort_by == 'oldest':
            profiles.sort(key=lambda p: (
                0 if p['issues'] else 1,
                min((i.reported_at or datetime.max for i in p['issues']),
                    default=datetime.max),
                name_key(p)))
        else:
            profiles.sort(key=name_key)
        return profiles

    lb_trucks = build_truck_profiles('LB')
    pli_trucks = build_truck_profiles('PLI')

    status_counts = {}
    for code, _label in ISSUE_STATUSES:
        status_counts[code] = (scoped_query(exclude='status')
                               .filter(Issue.current_status == code)
                               .count())

    severity_counts = {}
    for code, _label in ISSUE_SEVERITIES:
        severity_counts[code] = (scoped_query(exclude='severity')
                                 .filter(Issue.severity == code)
                                 .count())

    trucks = reportable_equipment_list()
    # Issues with no linked equipment appear in no truck panel — surface them
    # in their own section so they stay reachable.
    unlinked_issues = [i for i in all_issues if not i.equipment_id]

    resolved_issues = []
    resolved_count = 0
    if not status_filter:
        resolved_query = (scoped_query(exclude='status')
                          .filter(Issue.current_status.in_(RESOLVED_STATUSES)))
        resolved_count = resolved_query.count()
        # SQL-side (not Python) because of the .limit(20) — sorting after the
        # limit would only reorder the 20 newest
        if sort_by == 'oldest':
            resolved_order = (Issue.resolved_at.asc().nullslast(),
                              Issue.reported_at.asc())
        elif sort_by == 'severity':
            sev_rank = db.case(SEVERITY_RANK, value=Issue.severity, else_=0)
            resolved_order = (sev_rank.desc(),
                              Issue.resolved_at.desc().nullslast())
        else:  # 'unit' and 'recent': resolved most recently first
            resolved_order = (Issue.resolved_at.desc().nullslast(),
                              Issue.reported_at.desc())
        resolved_issues = (resolved_query
                           .options(joinedload(Issue.reporter),
                                    joinedload(Issue.equipment))
                           .order_by(*resolved_order)
                           .limit(20).all())

    return render_template('issues/queue.html',
                           lb_trucks=lb_trucks,
                           pli_trucks=pli_trucks,
                           lb_issue_count=sum(t['issue_count'] for t in lb_trucks),
                           pli_issue_count=sum(t['issue_count'] for t in pli_trucks),
                           unlinked_issues=unlinked_issues,
                           resolved_issues=resolved_issues,
                           resolved_count=resolved_count,
                           statuses=ISSUE_STATUSES,
                           severities=ISSUE_SEVERITIES,
                           categories=ISSUE_CATEGORIES,
                           trucks=trucks,
                           status_counts=status_counts,
                           severity_counts=severity_counts,
                           current_status=status_filter,
                           current_severity=severity_filter,
                           current_equipment=equipment_filter,
                           current_category=category_filter,
                           current_sort=sort_by,
                           search=search,
                           filters_active=filters_active)


# ── ISSUE DETAIL ──────────────────────────────────────────────────────

@issues_bp.route('/<int:issue_id>', methods=['GET'])
@shop_required
def detail(issue_id):
    issue = Issue.query.get_or_404(issue_id)
    history = issue.status_history.order_by(IssueStatusHistory.changed_at.asc()).all()
    return render_template('issues/detail.html',
                           issue=issue,
                           history=history,
                           statuses=ISSUE_STATUSES)


@issues_bp.route('/<int:issue_id>/status', methods=['POST'])
@shop_required
def update_status(issue_id):
    issue = Issue.query.get_or_404(issue_id)

    new_status = request.form.get('new_status', '')
    notes = request.form.get('notes', '').strip() or None

    # Optional return target (the queue's inline per-row form sends its own
    # filtered URL). Relative paths only — no open redirects.
    next_url = request.form.get('next', '')
    if not (next_url.startswith('/') and not next_url.startswith('//')):
        next_url = url_for('issues.detail', issue_id=issue.id)

    valid_statuses = [s[0] for s in ISSUE_STATUSES]
    if new_status not in valid_statuses:
        flash('Estado no válido.', 'danger')
        return redirect(next_url)

    if new_status == issue.current_status:
        flash('El estado no ha cambiado.', 'warning')
        return redirect(next_url)

    update_issue_status(issue, new_status, changed_by_user_id=current_user.id, notes=notes)
    db.session.commit()

    status_label = dict(ISSUE_STATUSES).get(new_status, new_status)
    flash(f'Estado actualizado a "{status_label}".', 'success')
    return redirect(next_url)


# ── ADMIN: GENERATE DRIVER LINK ──────────────────────────────────────

@issues_bp.route('/generate-link/<int:employee_id>', methods=['POST'])
@login_required
def generate_link(employee_id):
    if not current_user.is_admin:
        abort(403)
    emp = Employee.query.get_or_404(employee_id)
    emp.access_token = str(uuid.uuid4())
    db.session.commit()
    flash(f'Link de reporte generado para {emp.name}.', 'success')
    return redirect(url_for('employee_detail', id=employee_id))


# ── ADMIN: MANUAL ENTRY & BULK BACKLOG IMPORT ────────────────────────

def _coerce_date(cell):
    """Coerce an Excel cell (or date-string) to a date or None.

    Mirrors the equipment import's ``to_date`` including the 2-digit-year fix
    (Excel sometimes reads e.g. 2030 as 1930).
    """
    if cell is None or cell == '':
        return None
    if isinstance(cell, datetime):
        d = cell.date()
    elif isinstance(cell, date):
        d = cell
    elif isinstance(cell, str):
        s = cell.strip()
        if not s:
            return None
        d = None
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
            try:
                d = datetime.strptime(s, fmt).date()
                break
            except ValueError:
                continue
        if d is None:
            return None
    else:
        return None
    if d.year < 2000:
        try:
            d = d.replace(year=d.year + 100)
        except ValueError:
            pass
    return d


def _as_datetime(d):
    """A date (or None) → a datetime at midnight (or None)."""
    if d is None:
        return None
    return datetime(d.year, d.month, d.day)


def _is_future_date(d):
    """True if date/datetime d is after today (server date). None ⇒ False.

    Guards against absurd report/resolution dates (e.g. a spreadsheet's 2-digit
    year 35 bumped to 2035 by _coerce_date's +100 rule). Applied at every
    admin/import path that sets a date; the driver flow never sets one.
    """
    if d is None:
        return False
    day = d.date() if isinstance(d, datetime) else d
    return day > datetime.utcnow().date()


def _normalize_desc(text):
    """Lower-case and collapse whitespace for duplicate comparison."""
    return ' '.join((text or '').split()).lower()


def _canonical_report_day(reported_at):
    """The report DATE used for dedup, or None when the timestamp is volatile.

    Real spreadsheet dates are stored at midnight via _as_datetime(); a
    non-midnight time means "Fecha Reporte" was blank at import and a utcnow()
    fallback was used, so it must NOT anchor the dedup key (otherwise every
    re-import mints a new duplicate).
    """
    if not reported_at:
        return None
    if reported_at.hour or reported_at.minute or reported_at.second or reported_at.microsecond:
        return None
    return reported_at.date()


def find_duplicate_issue(equipment_id, category, description, report_day):
    """Return an existing Issue matching the import natural key, else None.

    Identity = (equipment_id, category, normalized description, canonical report
    day). ``report_day`` is the row's parsed date or None; a dateless row
    (None) matches a prior dateless issue regardless of which day each was
    imported (see _canonical_report_day). Compared in Python over the small
    per-truck+category candidate set so the normalization behaves identically
    on SQLite and Postgres.
    """
    norm = _normalize_desc(description)
    candidates = Issue.query.filter_by(equipment_id=equipment_id, category=category).all()
    for existing in candidates:
        if _normalize_desc(existing.description) != norm:
            continue
        if _canonical_report_day(existing.reported_at) == report_day:
            return existing
    return None


# ── SHOP MANAGER: MANUAL REPORT ENTRY ────────────────────────────────

# Cap on the "agregados en esta sesión" strip — keeps the save-and-add-another
# URL short while still showing a useful trail of what was just filed.
RECENT_CREATED_LIMIT = 10


def _created_ids(raw):
    """Parse the created-issue CSV carried through the save-and-add-another
    loop (same idiom as report_success's issue_ids)."""
    ids = [int(part) for part in (raw or '').split(',') if part.strip().isdigit()]
    return ids[-RECENT_CREATED_LIMIT:]


def _render_manual_form(form_data, status_code=200):
    """Single render path for the manual entry form. form_data is the source of
    sticky values: request.args on GET (query-arg prefill from the queue link
    and the save-and-add-another redirect), request.form on a 422 re-render."""
    created_ids = _created_ids(form_data.get('created', ''))
    created_issues = []
    if created_ids:
        found = {i.id: i for i in Issue.query.filter(Issue.id.in_(created_ids)).all()}
        # newest first, following the CSV's append order
        created_issues = [found[i] for i in reversed(created_ids) if i in found]
    return render_template('issues/manual_new.html',
                           trucks=reportable_equipment_list(),
                           categories=ISSUE_CATEGORIES,
                           severities=ISSUE_SEVERITIES,
                           statuses=ISSUE_STATUSES,
                           form_data=form_data,
                           created_issues=created_issues,
                           created_csv=','.join(str(i) for i in created_ids)), status_code


@issues_bp.route('/new', methods=['GET'])
@shop_manager_required
def new():
    return _render_manual_form(request.args)


@issues_bp.route('/new', methods=['POST'])
@shop_manager_required
def create():
    equipment_id = request.form.get('equipment_id', type=int)
    category = request.form.get('category', '')
    severity = request.form.get('severity', 'media')
    status = request.form.get('status', 'reportado') or 'reportado'
    description = request.form.get('description', '').strip()
    notes = request.form.get('notes', '').strip()

    errors = []
    if not equipment_id:
        errors.append('Debe seleccionar un camión.')
    elif not reportable_equipment_query().filter_by(id=equipment_id).first():
        errors.append('Camión no válido.')

    if category not in [c[0] for c in ISSUE_CATEGORIES]:
        errors.append('Categoría no válida.')
    if severity not in [s[0] for s in ISSUE_SEVERITIES]:
        errors.append('Severidad no válida.')
    if status not in [s[0] for s in ISSUE_STATUSES]:
        errors.append('Estado no válido.')
    if not description:
        errors.append('Debe incluir una descripción del problema.')

    report_date = _coerce_date(request.form.get('fecha_reporte', ''))
    resolved_date = _coerce_date(request.form.get('fecha_resuelto', ''))
    reported_at = _as_datetime(report_date) or datetime.utcnow()
    resolved_at = _as_datetime(resolved_date)

    if _is_future_date(report_date):
        errors.append('La fecha de reporte no puede estar en el futuro.')
    if _is_future_date(resolved_date):
        errors.append('La fecha de resolución no puede estar en el futuro.')

    if errors:
        for error in errors:
            flash(error, 'danger')
        return _render_manual_form(request.form, status_code=422)

    issue = Issue(
        equipment_id=equipment_id,
        category=category,
        severity=severity,
        description=description,
    )
    db.session.add(issue)
    db.session.flush()

    if current_user.is_admin:
        # Admins backfill the backlog: honor status, backdated dates, and note.
        initial_note = f'{BACKLOG_NOTE} — {notes}' if notes else BACKLOG_NOTE
        synthesize_status_history(issue, status, reported_at, resolved_at,
                                  initial_note=initial_note)
    else:
        # Shop managers add a live report: always a fresh 'reportado' dated now,
        # attributed to the manager's user id.
        update_issue_status(issue, 'reportado', changed_by_user_id=current_user.id,
                            notes='Reporte agregado manualmente por taller')
    db.session.commit()

    flash(f'Reporte #{issue.id} agregado.', 'success')

    if request.form.get('save_and_new'):
        # Rapid-entry loop: come straight back to a blank form with the truck
        # (and the admin's working date/status) carried forward.
        created = _created_ids(request.form.get('created', '')) + [issue.id]
        return redirect(url_for('issues.new',
                                equipment_id=equipment_id,
                                fecha_reporte=request.form.get('fecha_reporte') or None,
                                status=status if current_user.is_admin else None,
                                created=','.join(str(i) for i in created[-RECENT_CREATED_LIMIT:])))

    return redirect(url_for('issues.detail', issue_id=issue.id))


# ── ADMIN: QUICK-ADD REPORTABLE VEHICLE ──────────────────────────────

@issues_bp.route('/vehicle/new', methods=['GET'])
@admin_required
def vehicle_new():
    return render_template('issues/vehicle_new.html', companies=['LB', 'PLI'])


@issues_bp.route('/vehicle/new', methods=['POST'])
@admin_required
def vehicle_create():
    company = request.form.get('company', '')
    unit_number = request.form.get('unit_number', '').strip()
    plate_number = request.form.get('plate_number', '').strip()
    make = request.form.get('make', '').strip()
    model = request.form.get('model', '').strip()
    year_raw = request.form.get('year', '').strip()

    errors = []
    if company not in ('LB', 'PLI'):
        errors.append('Debe seleccionar una empresa (LB o PLI).')
    if not unit_number:
        errors.append('Debe incluir el número de unidad.')
    if model and (model.lower() in EXCLUDED_EQUIPMENT_MODELS
                  or classify_equipment(model) != 'truck'):
        errors.append(f'El modelo "{model}" no corresponde a un camión; '
                      'el equipo no aparecería en la lista de reportes.')

    from app import find_duplicate_equipment
    match = find_duplicate_equipment(None, plate_number, unit_number, company)
    if match:
        errors.append(f'Ya existe un camión con esa unidad/placa: {match.display_name} '
                      f'({match.company_full}). Verifique antes de crear un duplicado.')

    year = None
    if year_raw:
        try:
            year = int(year_raw)
        except ValueError:
            errors.append('El año no es válido.')

    if errors:
        for error in errors:
            flash(error, 'danger')
        return render_template('issues/vehicle_new.html',
                               companies=['LB', 'PLI']), 422

    equip = Equipment(
        company=company,
        equipment_type='vehicle',
        equipment_class=classify_equipment(model),
        status='activo',
        unit_number=unit_number,
        plate_number=plate_number,
        make=make,
        model=model,
        year=year,
    )
    db.session.add(equip)
    db.session.flush()

    # Create default permit slots (mirrors equipment_new in app.py).
    for code, name in EQUIPMENT_PERMIT_TYPES:
        if code == 'OTHER':
            continue
        applicability = 'YES'
        # Insurance is shared at the company level for LB/PLI.
        if code == 'SEGURO' and equip.company in ('LB', 'PLI'):
            applicability = 'N/A'
        db.session.add(EquipmentPermit(
            equipment_id=equip.id,
            permit_type=code,
            applicability=applicability,
        ))

    db.session.commit()
    flash(f'Camión {equip.display_name} agregado.', 'success')
    return redirect(url_for('issues.queue'))


@issues_bp.route('/import', methods=['GET'])
@admin_required
def import_form():
    session['issue_import_token'] = secrets.token_hex(16)
    return render_template('issues/import.html',
                           import_token=session['issue_import_token'],
                           categories=ISSUE_CATEGORIES,
                           severities=ISSUE_SEVERITIES,
                           statuses=ISSUE_STATUSES)


@issues_bp.route('/import', methods=['POST'])
@admin_required
@protect_uploaded_files()   # documents are manual-only; see models.py
def import_issues():
    from openpyxl import load_workbook
    from sqlalchemy.exc import IntegrityError

    # Idempotency: block double-submit / re-POST on refresh.
    token = request.form.get('import_token', '')
    if token != session.pop('issue_import_token', None):
        flash('Formulario ya fue procesado. Por favor intente nuevamente.', 'warning')
        return redirect(url_for('issues.import_form'))

    if 'file' not in request.files:
        flash('No se seleccionó archivo.', 'danger')
        return redirect(url_for('issues.import_form'))

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        flash('Solo se aceptan archivos Excel (.xlsx, .xls).', 'danger')
        return redirect(url_for('issues.import_form'))

    # Lookup maps: accept either the internal code or the Spanish label.
    cat_map = {c[0]: c[0] for c in ISSUE_CATEGORIES}
    cat_map.update({c[1].lower(): c[0] for c in ISSUE_CATEGORIES})
    sev_map = {s[0]: s[0] for s in ISSUE_SEVERITIES}
    sev_map.update({s[1].lower(): s[0] for s in ISSUE_SEVERITIES})
    status_map = {s[0]: s[0] for s in ISSUE_STATUSES}
    status_map.update({s[1].lower(): s[0] for s in ISSUE_STATUSES})

    # Truck lookup by unit_number, falling back to plate_number (both normalized).
    # Matches against ALL equipment so a historical backlog issue can attach to any
    # truck, including currently inactive/non-reportable ones. unit_number wins over
    # plate_number when both resolve to the same key.
    truck_by_unit = {}
    for eq in Equipment.query.all():
        if eq.unit_number:
            truck_by_unit[_normalize_unit(eq.unit_number)] = eq
    for eq in Equipment.query.all():
        if eq.plate_number:
            truck_by_unit.setdefault(_normalize_unit(eq.plate_number), eq)

    imported = 0
    duplicates = 0
    errors = []
    notices = []
    dup_notices = []
    seen = {}
    filepath = None
    try:
        fd, filepath = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        file.save(filepath)
        wb = load_workbook(filepath, read_only=True)

        if db.engine.dialect.name == 'postgresql':
            db.session.execute(db.text("SELECT pg_advisory_xact_lock(44)"))

        ws = None
        for name in wb.sheetnames:
            if 'reporte' in name.lower() or 'issue' in name.lower() or 'problema' in name.lower():
                ws = wb[name]
                break
        if ws is None:
            ws = wb[wb.sheetnames[0]]

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            wb.close()
            flash('El archivo está vacío.', 'warning')
            return redirect(url_for('issues.import_form'))

        # Reject the wrong template.
        headers = [str(h).strip().upper() if h else '' for h in rows[0]]
        normalized = [unicodedata.normalize('NFKD', h).encode('ascii', 'ignore').decode()
                      for h in headers]
        joined = ' '.join(normalized)
        if 'UNIDAD' not in joined or ('DESCRIPCION' not in joined and 'DESCRIPCIN' not in joined):
            wb.close()
            flash('El archivo no parece ser de reportes. Verifique que está usando la '
                  'plantilla correcta (Unidad, Categoría, Severidad, Descripción, '
                  'Estado, Fecha Reporte, Fecha Resuelto).', 'danger')
            return redirect(url_for('issues.import_form'))

        # Columns: Unidad | Categoría | Severidad | Descripción | Estado | Fecha Reporte | Fecha Resuelto
        for idx, row in enumerate(rows[1:], start=2):
            if not row or not any(row):
                continue

            def cell(i):
                return row[i] if i < len(row) else None

            unit_raw = cell(0)
            cat_raw = cell(1)
            sev_raw = cell(2)
            desc_raw = cell(3)
            status_raw = cell(4)
            reported_raw = cell(5)
            resolved_raw = cell(6)

            unit = _normalize_unit(unit_raw)
            truck = truck_by_unit.get(unit)
            if not truck:
                display = str(unit_raw).strip() if unit_raw is not None else ''
                errors.append(
                    f'Fila {idx}: no existe ninguna unidad con número/placa '
                    f'"{display or "(vacía)"}".')
                continue
            reason = _unreportable_reason(truck)

            description = str(desc_raw).strip() if desc_raw is not None else ''
            if not description:
                errors.append(f'Fila {idx}: falta la descripción.')
                continue

            category = cat_map.get(str(cat_raw).strip().lower()) if cat_raw is not None else None
            if not category:
                errors.append(f'Fila {idx}: categoría "{cat_raw}" no válida.')
                continue

            sev_key = str(sev_raw).strip().lower() if sev_raw is not None else ''
            severity = sev_map.get(sev_key, 'media') if sev_key else 'media'

            status_key = str(status_raw).strip().lower() if status_raw is not None else ''
            status = status_map.get(status_key, 'reportado') if status_key else 'reportado'
            if status_key and status_key not in status_map:
                errors.append(f'Fila {idx}: estado "{status_raw}" no válido.')
                continue

            report_date = _coerce_date(reported_raw)                      # date or None
            reported_at = _as_datetime(report_date) or datetime.utcnow()  # stored value
            resolved_date = _coerce_date(resolved_raw)
            resolved_at = _as_datetime(resolved_date)

            # Reject future dates (e.g. a 2-digit year 35 bumped to 2035) so the
            # bad row is counted as errored and skipped instead of persisted.
            if _is_future_date(report_date):
                errors.append(
                    f'Fila {idx}: la fecha de reporte ({report_date}) está en el futuro.')
                continue
            if _is_future_date(resolved_date):
                errors.append(
                    f'Fila {idx}: la fecha de resolución ({resolved_date}) está en el futuro.')
                continue

            # Skip duplicates so re-importing the same file is idempotent. Match
            # against rows already created in this batch (seen) and rows already
            # in the DB from a previous import (find_duplicate_issue). On a hit we
            # update the existing issue rather than create a new one. Key on the
            # PARSED date (None for dateless rows) — never the volatile utcnow
            # fallback — so dateless rows collapse instead of duplicating.
            dup_key = (truck.id, category, _normalize_desc(description),
                       report_date)
            existing = seen.get(dup_key) or find_duplicate_issue(
                truck.id, category, description, report_date)
            if existing:
                existing.severity = severity
                if status != existing.current_status:
                    synthesize_status_history(existing, status, reported_at, resolved_at)
                duplicates += 1
                dup_notices.append(
                    f'Fila {idx}: duplicado de "{truck.display_name}" — '
                    f'actualizado (no se creó uno nuevo).')
                continue

            issue = Issue(
                equipment_id=truck.id,
                category=category,
                severity=severity,
                description=description,
            )
            db.session.add(issue)
            db.session.flush()
            synthesize_status_history(issue, status, reported_at, resolved_at)
            seen[dup_key] = issue
            imported += 1

            if reason:
                notices.append(
                    f'Fila {idx}: unidad "{truck.display_name}" importada, pero está '
                    f'marcada con {reason} y no aparece en el formulario del chofer.')

        wb.close()
        db.session.commit()

    except IntegrityError:
        db.session.rollback()
        flash('Error: datos duplicados detectados. Posible importación simultánea.', 'danger')
        return redirect(url_for('issues.import_form'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error en la importación: {str(e)}', 'danger')
        return redirect(url_for('issues.import_form'))
    finally:
        if filepath and os.path.exists(filepath):
            os.remove(filepath)

    if imported or duplicates:
        flash(f'Importación completa: {imported} reportes importados, '
              f'{duplicates} duplicados omitidos, {len(errors)} con errores.', 'success')
    else:
        flash(f'No se importó ningún reporte. {len(errors)} fila(s) con errores.', 'warning')

    # Re-issue a token so the results page can be re-submitted from if needed.
    session['issue_import_token'] = secrets.token_hex(16)
    return render_template('issues/import.html',
                           import_token=session['issue_import_token'],
                           categories=ISSUE_CATEGORIES,
                           severities=ISSUE_SEVERITIES,
                           statuses=ISSUE_STATUSES,
                           import_errors=errors,
                           import_notices=notices,
                           dup_notices=dup_notices,
                           imported_count=imported,
                           duplicate_count=duplicates)
